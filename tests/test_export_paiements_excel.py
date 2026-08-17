"""Tests du rapport Excel des paiements : une ligne par facture, avec le
total versé et le détail des versements (date + montant) concaténés dans une
seule cellule — en-têtes fixes en chinois."""
from __future__ import annotations

from datetime import date

import openpyxl

from app.utils.export_paiements_excel import generer_rapport_paiements


def test_une_ligne_par_facture_avec_versements_concatenes(tmp_path):
    factures = [
        {"numero": 1, "client_nom": "ALPHA", "destination": "DAKAR", "total": 20000,
         "verse": 15000, "restant": 5000,
         "versements": [
             {"date_versement": date(2026, 7, 1), "montant": 10000},
             {"date_versement": date(2026, 7, 5), "montant": 5000},
         ]},
        {"numero": 2, "client_nom": "BETA", "destination": "THIES", "total": 10000,
         "verse": 0, "restant": 10000, "versements": []},
    ]
    chemin = generer_rapport_paiements(factures, tmp_path / "rapport.xlsx")
    assert chemin.exists()

    classeur = openpyxl.load_workbook(chemin)
    feuille = classeur.active
    assert feuille.title == "付款报表"
    lignes = list(feuille.iter_rows(values_only=True))
    assert lignes[0] == (
        "发票号", "客户", "目的地", "总额 (FCFA)", "已付总额 (FCFA)", "剩余 (FCFA)",
        "付款记录（日期: 金额）",
    )
    # une seule ligne pour la facture 1, malgré ses deux versements
    assert len(lignes) == 3
    assert lignes[1] == (1, "ALPHA", "DAKAR", 20000, 15000, 5000,
                         "01/07/2026: 10 000; 05/07/2026: 5 000")
    # aucun versement -> cellule vide, pas de ligne en moins
    assert lignes[2] == (2, "BETA", "THIES", 10000, 0, 10000, None)


def test_date_en_chaine_iso_acceptee(tmp_path):
    """Les réponses JSON de l'API sérialisent les dates en chaîne ISO."""
    factures = [
        {"numero": 1, "client_nom": "ALPHA", "destination": "DAKAR", "total": 5000,
         "verse": 5000, "restant": 0,
         "versements": [{"date_versement": "2026-07-01", "montant": 5000}]},
    ]
    chemin = generer_rapport_paiements(factures, tmp_path / "rapport2.xlsx")
    classeur = openpyxl.load_workbook(chemin)
    ligne = list(classeur.active.iter_rows(values_only=True))[1]
    assert ligne[6] == "01/07/2026: 5 000"
