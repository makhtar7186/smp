"""Tests : analyse produit, visuels de production et export Excel localisé."""
from __future__ import annotations

from datetime import date

import openpyxl

from app.models import LigneVente, Produit
from app.repositories.produit_repository import ProduitRepository
from app.services.export_excel_service import ExportExcelService
from app.services.facturation_service import FacturationService
from app.services.stats_service import StatsService


def _peupler(conn):
    produits = ProduitRepository(conn)
    produits.creer(Produit(nom="SM TAPISSIER", type_option="dimension",
                           valeur_option="180X190", prix=12000))
    produits.creer(Produit(nom="SM HOUSSE", type_option="dimension",
                           valeur_option="140X190", prix=10850))

    service = FacturationService(conn)
    # Mardi 14/07/2026 et samedi 18/07/2026
    service.enregistrer_facture(260001, date(2026, 7, 14), "CLIENT A", "DAKAR", [
        LigneVente(designation="SM TAPISSIER 180X190",
                   quantite=12, prix_unitaire=12000),
    ])
    service.enregistrer_facture(260002, date(2026, 7, 18), "CLIENT B", "BAKEL", [
        LigneVente(designation="SM TAPISSIER 180X190",
                   quantite=8, prix_unitaire=12000),
        LigneVente(designation="SM HOUSSE 140X190",
                   quantite=3, prix_unitaire=10850),
    ])
    return StatsService(conn)


DEBUT, FIN = date(2026, 1, 1), date(2026, 12, 31)


def test_resume_produit_specifique(conn):
    """Ex. demandé : nb de SM TAPISSIER 180X190 vendus sur une période."""
    stats = _peupler(conn)
    resume = stats.resume_produit("SM TAPISSIER 180X190", DEBUT, FIN)
    assert resume["quantite"] == 20            # 12 + 8
    assert resume["ca"] == 20 * 12000
    assert resume["nb_factures"] == 2
    assert resume["prix_moyen"] == 12000


def test_resume_produit_periode_restreinte(conn):
    stats = _peupler(conn)
    resume = stats.resume_produit("SM TAPISSIER 180X190",
                                  date(2026, 7, 15), date(2026, 7, 31))
    assert resume["quantite"] == 8


def test_liste_produits_vendus(conn):
    stats = _peupler(conn)
    libelles = stats.liste_produits_vendus()
    assert "SM TAPISSIER 180X190" in libelles
    assert "SM HOUSSE 140X190" in libelles


def test_serie_quantite_produit(conn):
    stats = _peupler(conn)
    serie = stats.serie_quantite_produit("SM TAPISSIER 180X190", DEBUT, FIN, "mois")
    assert serie == [("2026-07", 20)]


def test_quantites_par_dimension(conn):
    stats = _peupler(conn)
    resultat = dict(stats.quantites_par_dimension(DEBUT, FIN))
    assert resultat["180X190"] == 20
    assert resultat["140X190"] == 3


def test_quantites_par_litrage(conn):
    produits = ProduitRepository(conn)
    produits.creer(Produit(nom="BIDON", type_option="litrage", valeur_option="5L",
                           prix=2000))
    service = FacturationService(conn)
    service.enregistrer_facture(260010, date(2026, 7, 14), "CLIENT C", "DAKAR", [
        LigneVente(designation="BIDON 5L", quantite=6, prix_unitaire=2000),
    ])
    resultat = dict(StatsService(conn).quantites_par_litrage(DEBUT, FIN))
    assert resultat["5L"] == 6


def test_ca_par_jour_semaine(conn):
    stats = _peupler(conn)
    resultat = dict(stats.ca_par_jour_semaine(DEBUT, FIN))
    assert resultat[2] == 12 * 12000               # mardi 14/07/2026
    assert resultat[6] == 8 * 12000 + 3 * 10850    # samedi 18/07/2026


def test_top_clients(conn):
    stats = _peupler(conn)
    top = stats.top_clients(DEBUT, FIN)
    assert top[0][0] == "CLIENT A"
    assert len(top) == 2


def test_export_excel_entetes_selon_langue(conn, tmp_path):
    from app.i18n import translations
    _peupler(conn)
    service = ExportExcelService(conn, dossier_exports=tmp_path)
    try:
        translations.definir_langue("zh")
        chemin = service.exporter_historique()
        feuille = openpyxl.load_workbook(chemin)["历史"]
        assert feuille.cell(1, 2).value == "客户"
        assert feuille.cell(1, 7).value == "数量"
        translations.definir_langue("en")
        chemin = service.exporter_historique()
        feuille = openpyxl.load_workbook(chemin)["History"]
        assert feuille.cell(1, 2).value == "Customer"
    finally:
        translations.definir_langue("fr")
