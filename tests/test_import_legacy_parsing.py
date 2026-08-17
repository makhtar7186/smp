"""Tests de `app.client.ui.analyser_feuille_import_legacy` — lecture d'un
résumé de paiements historiques Excel, colonnes lues par position (jamais par
en-tête, potentiellement en chinois). Voir CLAUDE.md, section « Import des
paiements historiques »."""
from __future__ import annotations

import openpyxl
import pytest

from app.client.ui import analyser_feuille_import_legacy


def _classeur(lignes_donnees: list[list]):
    """Construit un classeur avec 3 feuilles vides puis une 4ᵉ portant le
    résumé (comme le fichier source réel), 3 lignes sautées puis l'en-tête à
    la ligne 4 (ici en chinois, jamais lu par le code)."""
    wb = openpyxl.Workbook()
    wb.active.title = "Feuille1"
    wb.create_sheet("Feuille2")
    wb.create_sheet("Feuille3")
    feuille = wb.create_sheet("表4")
    feuille.append(["标题", None, None])
    feuille.append([None])
    feuille.append([None])
    feuille.append(["日期", "单号", "客户", "", "", "", "金额", "金额", "已付款", "余额", "备注"])
    for ligne in lignes_donnees:
        feuille.append(ligne)
    return wb, feuille


class TestNominal:
    def test_ligne_complete_valide(self):
        wb, feuille = _classeur([
            ["2024-01-15", 1, "ANCIEN CLIENT 001", None, None, None,
             50000, 50000, 30000, 20000, "commentaire"],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert analyse.lignes_ignorees == []
        assert analyse.avertissements == []
        assert len(analyse.lignes_valides) == 1
        ligne = analyse.lignes_valides[0]
        assert ligne == {
            "numero": 1, "date_facture": "2024-01-15", "client_nom": "ANCIEN CLIENT 001",
            "montant": 50000, "versement": 30000, "commentaire": "commentaire",
        }

    def test_date_en_texte_jj_mm_aaaa(self):
        wb, feuille = _classeur([
            ["15/01/2024", 2, "CLIENT", None, None, None, 1000, 1000, 0, 1000, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert analyse.lignes_valides[0]["date_facture"] == "2024-01-15"

    def test_montant_texte_avec_espaces(self):
        wb, feuille = _classeur([
            ["2024-01-15", 3, "CLIENT", None, None, None, "50 000", "50 000", 0, 50000, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert analyse.lignes_valides[0]["montant"] == 50000

    def test_arret_a_la_premiere_ligne_vide(self):
        wb, feuille = _classeur([
            ["2024-01-15", 1, "CLIENT A", None, None, None, 1000, 1000, 0, 1000, ""],
            [None, None, None, None, None, None, None, None, None, None, None],
            ["2024-01-16", 2, "CLIENT B", None, None, None, 2000, 2000, 0, 2000, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert len(analyse.lignes_valides) == 1
        assert analyse.lignes_valides[0]["numero"] == 1


class TestLignesIgnoreesLocalement:
    def test_numero_manquant(self):
        wb, feuille = _classeur([
            ["2024-01-15", None, "CLIENT", None, None, None, 1000, 1000, 0, 1000, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert analyse.lignes_valides == []
        assert len(analyse.lignes_ignorees) == 1

    def test_date_illisible(self):
        wb, feuille = _classeur([
            ["pas une date", 1, "CLIENT", None, None, None, 1000, 1000, 0, 1000, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert analyse.lignes_valides == []

    def test_client_vide(self):
        wb, feuille = _classeur([
            ["2024-01-15", 1, "", None, None, None, 1000, 1000, 0, 1000, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert analyse.lignes_valides == []

    def test_montant_nul(self):
        wb, feuille = _classeur([
            ["2024-01-15", 1, "CLIENT", None, None, None, 0, 0, 0, 0, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert analyse.lignes_valides == []


class TestAvertissements:
    def test_montants_colonne7_et_8_differents(self):
        wb, feuille = _classeur([
            ["2024-01-15", 1, "CLIENT", None, None, None, 1000, 1200, 0, 1000, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert len(analyse.lignes_valides) == 1
        assert analyse.lignes_valides[0]["montant"] == 1000  # colonne 7 préférée
        assert len(analyse.avertissements) == 1

    def test_montant_colonne7_vide_repli_colonne8(self):
        wb, feuille = _classeur([
            ["2024-01-15", 1, "CLIENT", None, None, None, None, 1500, 0, 1500, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert analyse.lignes_valides[0]["montant"] == 1500
        assert analyse.avertissements == []

    def test_solde_incoherent(self):
        wb, feuille = _classeur([
            ["2024-01-15", 1, "CLIENT", None, None, None, 1000, 1000, 200, 999, ""],
        ])
        analyse = analyser_feuille_import_legacy(feuille)
        assert len(analyse.lignes_valides) == 1  # importée quand même
        assert len(analyse.avertissements) == 1
