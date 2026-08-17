"""Tests de l'export Excel de l'historique des ventes."""
from __future__ import annotations

from datetime import date

import openpyxl

from app.models import LigneVente
from app.services.export_excel_service import ExportExcelService
from app.services.facturation_service import FacturationService


def test_export_historique_complet(conn, tmp_path):
    service = FacturationService(conn)
    service.enregistrer_facture(260001, date(2026, 6, 15), "CLIENT A", "DAKAR", [
        LigneVente(designation="SM TAPISSIER 90X190X", 
                   quantite=10, prix_unitaire=14130),
        LigneVente(designation="SM HOUSSE 140X190X", 
                   quantite=4, prix_unitaire=10850),
    ])
    service.enregistrer_facture(260002, date(2026, 7, 18), "CLIENT B", "BAKEL", [
        LigneVente(designation="SM RESSORT 180X190X", 
                   quantite=1, prix_unitaire=78000),
    ])

    chemin = ExportExcelService(conn, dossier_exports=tmp_path).exporter_historique()
    assert chemin.exists()

    feuille = openpyxl.load_workbook(chemin)["Historique"]
    assert feuille.max_row == 4                       # en-tête + 3 lignes
    assert feuille.cell(1, 2).value == "Client"
    assert feuille.cell(2, 2).value == "CLIENT A"
    # openpyxl relit les dates en datetime
    assert feuille.cell(2, 3).value.date() == date(2026, 6, 15)
    assert feuille.cell(2, 3).number_format == "DD/MM/YYYY"
    assert feuille.cell(2, 9).value == 141300        # total ligne calculé
    assert feuille.cell(4, 4).value == 260002


def test_export_base_vide(conn, tmp_path):
    chemin = ExportExcelService(conn, dossier_exports=tmp_path).exporter_historique()
    feuille = openpyxl.load_workbook(chemin)["Historique"]
    assert feuille.max_row == 1                       # uniquement l'en-tête
