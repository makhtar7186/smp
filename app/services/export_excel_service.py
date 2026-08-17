"""Export Excel de l'historique complet des ventes."""
from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import config
from app.i18n.translations import t
from app.utils.export_paiements_excel import generer_rapport_paiements

# Clés i18n des en-têtes : traduites dans la langue active au moment de l'export
_CLES_ENTETES = ["exp_no_ligne", "fact_client", "date", "fact_numero",
                 "fact_destination", "fact_designation",
                 "fact_quantite", "fact_prix_unitaire", "fact_prix_total",
                 "exp_remarque"]
_LARGEURS = [10, 26, 12, 14, 18, 30, 10, 14, 14, 20]


class ExportExcelService:
    """Produit un classeur .xlsx contenant toutes les lignes de vente."""

    def __init__(self, conn: sqlite3.Connection,
                 dossier_exports: Path | None = None) -> None:
        self._conn = conn
        self._dossier = dossier_exports or config.DOSSIER_EXPORTS
        self._dossier.mkdir(parents=True, exist_ok=True)

    def exporter_historique(self) -> Path:
        """Écrit l'historique complet (toutes les lignes de vente) en Excel."""
        rows = self._conn.execute(
            "SELECT l.id AS ligne_id, f.client_nom, f.date_facture, f.numero,"
            " f.destination, l.designation, l.quantite,"
            " l.prix_unitaire, l.quantite * l.prix_unitaire AS total, l.remarque"
            " FROM lignes_vente l JOIN factures f ON f.id = l.facture_id"
            " ORDER BY f.date_facture, f.numero, l.id"
        ).fetchall()

        classeur = openpyxl.Workbook()
        feuille = classeur.active
        feuille.title = t("exp_feuille")

        remplissage = PatternFill("solid", fgColor="1D4ED8")
        for colonne, cle in enumerate(_CLES_ENTETES, start=1):
            cellule = feuille.cell(row=1, column=colonne, value=t(cle))
            cellule.font = Font(bold=True, color="FFFFFF")
            cellule.fill = remplissage
            cellule.alignment = Alignment(horizontal="center")
            feuille.column_dimensions[get_column_letter(colonne)].width = \
                _LARGEURS[colonne - 1]

        for rang, row in enumerate(rows, start=2):
            date_cellule = feuille.cell(
                row=rang, column=3,
                value=datetime.strptime(row["date_facture"], "%Y-%m-%d").date(),
            )
            date_cellule.number_format = "DD/MM/YYYY"
            valeurs = {
                1: row["ligne_id"], 2: row["client_nom"], 4: row["numero"],
                5: row["destination"], 6: row["designation"],
                7: row["quantite"],
                8: row["prix_unitaire"], 9: row["total"],
                10: row["remarque"],
            }
            for colonne, valeur in valeurs.items():
                cellule = feuille.cell(row=rang, column=colonne, value=valeur)
                if colonne in (8, 9):
                    cellule.number_format = "#,##0"

        feuille.freeze_panes = "A2"
        horodatage = datetime.now().strftime("%Y%m%d_%H%M")
        chemin = self._dossier / f"HISTORIQUE_VENTES_{horodatage}.xlsx"
        classeur.save(chemin)
        return chemin

    def exporter_paiements(self, lignes: list[dict]) -> Path:
        """Rapport de paiement (page Paiements) — en-têtes fixes en chinois,
        volontairement indépendantes de la langue active de l'application
        (demande explicite, contrairement aux autres exports de ce service).
        `lignes` : dicts numero/client_nom/destination/total/verse/restant,
        exactement l'ensemble filtré affiché sur la page. Logique de
        génération partagée avec le mode client, voir
        `app.utils.export_paiements_excel`."""
        horodatage = datetime.now().strftime("%Y%m%d_%H%M")
        chemin = self._dossier / f"RAPPORT_PAIEMENTS_{horodatage}.xlsx"
        return generer_rapport_paiements(lignes, chemin)
