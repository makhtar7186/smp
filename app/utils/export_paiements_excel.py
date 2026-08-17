"""Génère le rapport de paiement au format Excel — en-têtes fixes en
**chinois**, volontairement indépendantes de la langue active de
l'application (demande explicite), contrairement aux autres exports qui
suivent la langue de l'UI (`ExportExcelService`). Ce module ne dépend que
d'openpyxl et de la stdlib (pas de SQLite/repository) afin d'être utilisable
tel quel par le mode client (`app/client/ui.py`), qui n'a pas de base
locale, en plus de l'application principale (`ExportExcelService.exporter_paiements`).

Une ligne par **facture** (pas par versement) : le total versé et le
détail des versements (date + montant, concaténés dans une seule cellule)
apparaissent sur la même ligne — une ligne par versement mélangerait
numéro/client/destination répétés sur plusieurs lignes pour une même
facture, ce qui prête à confusion."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ENTETES = ["发票号", "客户", "目的地", "总额 (FCFA)", "已付总额 (FCFA)", "剩余 (FCFA)",
            "付款记录（日期: 金额）"]
_LARGEURS = [12, 26, 22, 16, 16, 16, 40]


def _vers_date(valeur: date | str) -> date:
    """Accepte un `date` déjà construit ou une chaîne ISO (YYYY-MM-DD, format
    des réponses JSON de l'API) et retourne toujours un `date`."""
    if isinstance(valeur, str):
        return datetime.strptime(valeur[:10], "%Y-%m-%d").date()
    return valeur


def _formatter_montant(montant: int) -> str:
    return f"{montant:,}".replace(",", " ")


def _formatter_versements(versements: list[dict]) -> str:
    """Concatène chaque versement en « JJ/MM/AAAA: montant », séparés par
    « ; », du plus ancien au plus récent — une seule cellule, pas une ligne
    par versement."""
    parties = []
    for versement in versements:
        jour = _vers_date(versement["date_versement"]).strftime("%d/%m/%Y")
        parties.append(f"{jour}: {_formatter_montant(versement['montant'])}")
    return "; ".join(parties)


def generer_rapport_paiements(factures: list[dict], chemin: Path) -> Path:
    """Écrit le rapport à `chemin` (créé/écrasé) et le retourne.

    `factures` : dicts avec les clés numero/client_nom/destination/total/
    verse/restant (format déjà utilisé par `PaiementService.lister_avec_solde`
    et par les réponses JSON de `GET /paiements/factures`), plus une clé
    `versements` : liste de dicts `date_versement`/`montant` (le format de
    `VersementOut`/`Versement`, du plus ancien au plus récent) — absente ou
    vide si la facture n'a encore reçu aucun versement.
    """
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "付款报表"

    remplissage = PatternFill("solid", fgColor="1D4ED8")
    for colonne, entete in enumerate(_ENTETES, start=1):
        cellule = feuille.cell(row=1, column=colonne, value=entete)
        cellule.font = Font(bold=True, color="FFFFFF")
        cellule.fill = remplissage
        cellule.alignment = Alignment(horizontal="center")
        feuille.column_dimensions[get_column_letter(colonne)].width = _LARGEURS[colonne - 1]

    for rang, facture in enumerate(factures, start=2):
        feuille.cell(row=rang, column=1, value=facture["numero"])
        feuille.cell(row=rang, column=2, value=facture["client_nom"])
        feuille.cell(row=rang, column=3, value=facture.get("destination", ""))
        for colonne, cle in ((4, "total"), (5, "verse"), (6, "restant")):
            cellule = feuille.cell(row=rang, column=colonne, value=facture[cle])
            cellule.number_format = "#,##0"
        feuille.cell(row=rang, column=7,
                     value=_formatter_versements(facture.get("versements") or []))

    feuille.freeze_panes = "A2"
    chemin.parent.mkdir(parents=True, exist_ok=True)
    classeur.save(chemin)
    return chemin
