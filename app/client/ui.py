"""Interface Tkinter du mode client : historique + impression des factures
usine (existant), gestion des paiements (versements — écriture réservée au
rôle `role_client`) et vue client (synthèse + remise **annuelle**, gérée par
ce poste ; la remise **par facture** reste réservée à l'application
principale, qui écrit les factures — voir CLAUDE.md, section « Paiements »).
Interface trilingue fr/中文 (le mode client n'expose pas l'anglais), langue
persistée dans `client_config.json` — jamais dans les réglages de
l'application principale (`app/data/settings.json`), que ce poste n'a pas
vocation à lire.

**Accès aux données** : `self.api` est soit un `ApiClient` (HTTP, poste
observateur d'une machine réellement distante), soit un `AccesDirectDonnees`
(`app/client/acces_direct.py`, lectures directes en SQLite) quand ce poste
héberge lui-même sa base (`_config_auto_locale`/`_api_direct` ci-dessous) —
les deux exposent la même interface, les onglets ne savent jamais lequel ils
utilisent. Les écritures (versement, remise annuelle) passent **toujours**
par HTTP local dans les deux cas, pour préserver la file FIFO du serveur
(`app/api/write_queue.py`) qui protège contre une écriture concurrente d'un
poste réellement distant.

**Onglet Serveur** (`OngletServeur`) : quand ce poste héberge lui-même la base
+ le serveur API (cas de la machine du boss, dans la topologie inversée où
« SMP.exe » n'est plus nécessaire — voir CLAUDE.md, section « Machine
de facturation »), ce poste pilote directement `SMP-Serveur.exe`
(processus séparé, copié dans le même dossier) via `ApiManagementService` —
exactement le même service que la page « API distante » de l'application
principale, mais intégré ici pour que ce poste n'ait besoin d'installer
qu'un seul exécutable. `ApiManagementService` ne dépend que de `requests` et
du module standard `subprocess` : aucune dépendance FastAPI/uvicorn ajoutée
à `SMP-Client.exe`."""
from __future__ import annotations

import tempfile
import tkinter as tk
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk

import openpyxl

from app import config
from app.client import archive_client, queue_hors_ligne
from app.client.acces_direct import AccesDirectDonnees
from app.client.api_client import (
    ApiClient, ErreurApiClient, ErreurAuthentification, ErreurInjoignable,
    ErreurPermission,
)
from app.client.config_client import ConfigClient, charger, sauvegarder
from app.client import config_postgres
from app.i18n import translations
from app.i18n.translations import t
from app.repositories.base_repository import creer_connexion
from app.services.api_management_service import ApiManagementService, ErreurGestionApi
from app.services.postgres_migration_service import (
    ConfigPostgres, PostgresMigrationService,
)
from app.ui import theme
from app.ui.components import graphiques_dashboard as graph
from app.ui.components.champ_date import ChampDate
from app.ui.components.data_table import DataTable
from app.ui.components.kpi_card import KPICard
from app.ui.components.toast import afficher_toast
from app.utils.export_paiements_excel import generer_rapport_paiements
from app.utils.fichiers import ouvrir_fichier
from app.utils.formatting import (
    format_date, format_fcfa, format_nombre, parse_date_affichage,
)
from app.utils.recherche import correspond, decouper_termes

_LANGUES_CLIENT = ("fr", "zh")  # sous-ensemble volontaire (pas d'anglais côté client)


class FenetreConfig(tk.Toplevel):
    """Boîte de dialogue de saisie/édition de la connexion (IP, port, jeton,
    identifiant de poste optionnel)."""

    def __init__(self, parent: tk.Tk, config: ConfigClient) -> None:
        super().__init__(parent)
        self.title(t("client_connexion_titre"))
        self.resizable(False, False)
        self.configure(bg=theme.COULEURS["fond"])
        self.resultat: ConfigClient | None = None
        self.transient(parent)
        self.grab_set()

        cadre = ttk.Frame(self, padding=theme.PAD_L)
        cadre.pack(fill="both", expand=True)

        ttk.Label(cadre, text=t("client_hote"), style="SousTitre.TLabel").pack(anchor="w")
        self._champ_hote = ttk.Entry(cadre)
        self._champ_hote.insert(0, config.hote)
        self._champ_hote.pack(fill="x", pady=(2, 12))

        ttk.Label(cadre, text=t("client_port"), style="SousTitre.TLabel").pack(anchor="w")
        self._champ_port = ttk.Entry(cadre)
        self._champ_port.insert(0, str(config.port))
        self._champ_port.pack(fill="x", pady=(2, 12))

        ttk.Label(cadre, text=t("client_jeton"), style="SousTitre.TLabel").pack(anchor="w")
        self._champ_token = ttk.Entry(cadre, show="•")
        self._champ_token.insert(0, config.token)
        self._champ_token.pack(fill="x", pady=(2, 12))

        ttk.Label(cadre, text=t("client_machine_id"), style="SousTitre.TLabel").pack(anchor="w")
        self._champ_machine = ttk.Entry(cadre)
        self._champ_machine.insert(0, config.machine_id)
        self._champ_machine.pack(fill="x", pady=(2, 16))

        boutons = ttk.Frame(cadre)
        boutons.pack(fill="x")
        ttk.Button(boutons, text=t("annuler"), command=self.destroy).pack(side="right")
        ttk.Button(boutons, text=t("client_enregistrer_connexion"), style="Accent.TButton",
                  command=self._valider).pack(side="right", padx=(0, 8))

        # Taille calculée d'après le contenu réel (voir FenetreConfigSync,
        # app/ui/app_facturation.py, pour le même correctif).
        self.update_idletasks()
        self.geometry(f"{max(420, self.winfo_reqwidth())}x{self.winfo_reqheight()}")

    def _valider(self) -> None:
        hote = self._champ_hote.get().strip()
        token = self._champ_token.get().strip()
        machine_id = self._champ_machine.get().strip()
        try:
            port = int(self._champ_port.get().strip())
        except ValueError:
            messagebox.showerror(t("erreur"), t("client_port_invalide"), parent=self)
            return
        if not hote or not token:
            messagebox.showerror(t("erreur"), t("client_hote_jeton_requis"), parent=self)
            return
        self.resultat = ConfigClient(hote=hote, port=port, token=token,
                                     machine_id=machine_id)
        self.destroy()


class DialogueVersement(simpledialog.Dialog):
    """Saisie du montant, de la date et d'une remarque optionnelle (nature du
    versement, ex. « acompte », « règlement final ») pour un versement."""

    def __init__(self, parent: tk.Widget) -> None:
        self.resultat: tuple[int, date, str] | None = None
        super().__init__(parent, title=t("client_versement_titre"))

    def body(self, master: tk.Widget) -> tk.Widget:
        ttk.Label(master, text=t("client_versement_montant")).grid(
            row=0, column=0, sticky="w", pady=4)
        self._champ_montant = ttk.Entry(master)
        self._champ_montant.grid(row=0, column=1, pady=4)
        ttk.Label(master, text=t("client_versement_date")).grid(
            row=1, column=0, sticky="w", pady=4)
        self._champ_date = ChampDate(master)
        self._champ_date.grid(row=1, column=1, pady=4)
        ttk.Label(master, text=t("client_versement_remarque")).grid(
            row=2, column=0, sticky="w", pady=4)
        self._champ_remarque = ttk.Entry(master, width=28)
        self._champ_remarque.grid(row=2, column=1, pady=4)
        return self._champ_montant

    def validate(self) -> bool:
        try:
            montant = int(self._champ_montant.get().strip().replace(" ", ""))
            date_versement = parse_date_affichage(self._champ_date.get().strip())
        except ValueError:
            messagebox.showerror(t("erreur"), t("client_montant_date_invalide"), parent=self)
            return False
        if montant == 0:
            messagebox.showerror(t("erreur"), t("paie_montant_nul"), parent=self)
            return False
        self.resultat = (montant, date_versement, self._champ_remarque.get().strip())
        return True


class DialogueRemiseAnnuelle(simpledialog.Dialog):
    """Saisie année / CA annuel / taux / note pour la remise annuelle d'un client."""

    def __init__(self, parent: tk.Widget, annee_defaut: int, ca_defaut: int) -> None:
        self._annee_defaut = annee_defaut
        self._ca_defaut = ca_defaut
        self.resultat: tuple[int, int, float, str] | None = None
        super().__init__(parent, title=t("client_remise_annuelle_titre"))

    def body(self, master: tk.Widget) -> tk.Widget:
        ttk.Label(master, text=t("rem_annee")).grid(row=0, column=0, sticky="w", pady=4)
        self._champ_annee = ttk.Entry(master)
        self._champ_annee.insert(0, str(self._annee_defaut))
        self._champ_annee.grid(row=0, column=1, pady=4)
        ttk.Label(master, text=t("rem_ca_annuel")).grid(row=1, column=0, sticky="w", pady=4)
        self._champ_ca = ttk.Entry(master)
        self._champ_ca.insert(0, str(self._ca_defaut))
        self._champ_ca.grid(row=1, column=1, pady=4)
        ttk.Label(master, text=t("rem_taux")).grid(row=2, column=0, sticky="w", pady=4)
        self._champ_taux = ttk.Entry(master)
        self._champ_taux.grid(row=2, column=1, pady=4)
        ttk.Label(master, text=t("rem_note")).grid(row=3, column=0, sticky="w", pady=4)
        self._champ_note = ttk.Entry(master)
        self._champ_note.grid(row=3, column=1, pady=4)
        return self._champ_annee

    def validate(self) -> bool:
        try:
            annee = int(self._champ_annee.get().strip())
            ca_annuel = int(self._champ_ca.get().strip().replace(" ", ""))
            taux = float(self._champ_taux.get().strip().replace(",", "."))
            if not 0 <= taux <= 100:
                raise ValueError
        except ValueError:
            messagebox.showerror(t("erreur"), t("nombre_invalide"), parent=self)
            return False
        self.resultat = (annee, ca_annuel, taux, self._champ_note.get().strip())
        return True


class OngletHistorique(ttk.Frame):
    """Historique des factures usine + impression (comportement inchangé),
    avec en plus le total facturé de l'ensemble filtré affiché en bas."""

    def __init__(self, parent: tk.Widget, app: "ApplicationClient") -> None:
        super().__init__(parent)
        self._app = app
        self._factures_brutes: list[dict] = []
        self._factures: dict[int, dict] = {}
        self._construire()

    def _construire(self) -> None:
        filtres = ttk.Frame(self, padding=(theme.PAD, theme.PAD))
        filtres.pack(fill="x")
        ttk.Label(filtres, text=t("du")).pack(side="left")
        self._champ_du = ChampDate(filtres, largeur=12)
        self._champ_du.pack(side="left", padx=(4, 12))
        self._champ_du.vider()
        ttk.Label(filtres, text=t("au")).pack(side="left")
        self._champ_au = ChampDate(filtres, largeur=12)
        self._champ_au.pack(side="left", padx=(4, 12))
        self._champ_au.vider()
        ttk.Button(filtres, text=t("filtrer"), command=self.actualiser).pack(side="left")
        ttk.Button(filtres, text=t("reinitialiser"), command=self._reinitialiser_filtres).pack(
            side="left", padx=(6, 0))

        recherche = ttk.Frame(self, padding=(theme.PAD, 0))
        recherche.pack(fill="x")
        ttk.Label(recherche, text=t("client_recherche_client_destination")).pack(side="left")
        self.var_recherche_client = tk.StringVar()
        self.var_recherche_client.trace_add("write", lambda *_: self._appliquer_filtres_locaux())
        ttk.Entry(recherche, textvariable=self.var_recherche_client, width=36).pack(
            side="left", padx=(4, 12))
        ttk.Label(recherche, text=t("client_facture_numero")).pack(side="left")
        self.var_recherche_numero = tk.StringVar()
        self.var_recherche_numero.trace_add("write", lambda *_: self._appliquer_filtres_locaux())
        ttk.Entry(recherche, textvariable=self.var_recherche_numero, width=12).pack(
            side="left", padx=(4, 12))
        ttk.Label(recherche, text=t("fact_telephone")).pack(side="left")
        self.var_recherche_telephone = tk.StringVar()
        self.var_recherche_telephone.trace_add("write", lambda *_: self._appliquer_filtres_locaux())
        ttk.Entry(recherche, textvariable=self.var_recherche_telephone, width=16).pack(
            side="left", padx=(4, 0))
        self.var_inclure_archivees = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            recherche, text=t("client_inclure_archivees"),
            variable=self.var_inclure_archivees,
            command=self._appliquer_filtres_locaux,
        ).pack(side="left", padx=(theme.PAD_L, 0))

        self._statut = ttk.Label(self, text="", style="Secondaire.TLabel",
                                 padding=(theme.PAD, 4))
        self._statut.pack(fill="x")

        corps = ttk.Frame(self, padding=theme.PAD)
        corps.pack(fill="both", expand=True)

        colonnes = ("numero", "date", "client", "destination", "telephone", "total")
        self._table = ttk.Treeview(corps, columns=colonnes, show="headings",
                                   selectmode="extended")
        titres = {"numero": t("fact_no_article"), "date": t("date"),
                 "client": t("fact_client"), "destination": t("fact_destination"),
                 "telephone": t("fact_telephone"), "total": t("total")}
        for cle in colonnes:
            self._table.heading(cle, text=titres[cle])
        self._table.column("numero", width=90, anchor="center")
        self._table.column("date", width=100, anchor="center")
        self._table.column("client", width=220)
        self._table.column("destination", width=170)
        self._table.column("telephone", width=110, anchor="center")
        self._table.column("total", width=140, anchor="e")
        theme.configurer_lignes_alternees(self._table)
        self._table.tag_configure("archivee", foreground=theme.COULEURS["texte_secondaire"])
        self._table.pack(fill="both", expand=True, side="left")

        defilement = ttk.Scrollbar(corps, orient="vertical", command=self._table.yview)
        self._table.configure(yscrollcommand=defilement.set)
        defilement.pack(fill="y", side="right")

        pied = ttk.Frame(self, padding=theme.PAD)
        pied.pack(fill="x")
        ttk.Button(pied, text=t("client_imprimer_facture"), style="Accent.TButton",
                  command=self._imprimer_facture).pack(side="left")
        ttk.Button(pied, text=t("client_bordereau_livraison"),
                  command=self._imprimer_bordereau).pack(side="left", padx=(8, 0))
        ttk.Button(pied, text=t("client_tout_selectionner"),
                  command=self._tout_selectionner).pack(side="left", padx=(8, 0))
        ttk.Button(pied, text=t("client_archiver_selection"),
                  command=self._archiver_selection).pack(side="left", padx=(8, 0))
        ttk.Button(pied, text=t("client_desarchiver_selection"),
                  command=self._desarchiver_selection).pack(side="left", padx=(8, 0))
        self._label_total_somme = ttk.Label(
            pied, text="", font=theme.POLICES["total_gros"],
            foreground=theme.COULEURS["accent_sombre"])
        self._label_total_somme.pack(side="right")

    def _reinitialiser_filtres(self) -> None:
        self._champ_du.delete(0, "end")
        self._champ_au.delete(0, "end")
        self.var_recherche_client.set("")
        self.var_recherche_numero.set("")
        self.var_recherche_telephone.set("")
        self.actualiser()

    def _parse_champ_date(self, champ: ttk.Entry) -> date | None:
        texte = champ.get().strip()
        if not texte:
            return None
        try:
            return parse_date_affichage(texte)
        except ValueError:
            messagebox.showerror(t("erreur"), t("date_invalide"))
            return None

    def actualiser(self) -> None:
        api = self._app.api
        if api is None:
            return
        date_debut = self._parse_champ_date(self._champ_du)
        date_fin = self._parse_champ_date(self._champ_au)
        try:
            factures = api.lister_factures(date_debut, date_fin)
        except ErreurInjoignable:
            self._statut.configure(text=t("client_injoignable"), foreground=theme.COULEURS["danger"])
            return
        except ErreurAuthentification:
            self._statut.configure(
                text=t("client_jeton_refuse"), foreground=theme.COULEURS["danger"])
            return
        except ErreurApiClient as exc:
            self._statut.configure(text=str(exc), foreground=theme.COULEURS["danger"])
            return

        self._factures_brutes = factures
        self._appliquer_filtres_locaux()

    def _appliquer_filtres_locaux(self) -> None:
        """Filtre par client/destination, par numéro et par téléphone du
        client (recherche instantanée, côté client) la dernière liste
        récupérée de l'API — évite un aller-retour réseau à chaque frappe,
        contrairement aux dates (filtrées côté serveur). Chaque champ
        accepte plusieurs valeurs séparées par une virgule (logique OU, ex.
        « ALPHA, BETA »). Filtre aussi par l'archivage **local** au poste
        (indépendant de celui de l'application principale — voir
        `app.client.archive_client`)."""
        termes_client = decouper_termes(self.var_recherche_client.get())
        termes_numero = decouper_termes(self.var_recherche_numero.get())
        termes_telephone = decouper_termes(self.var_recherche_telephone.get())
        archivees = archive_client.charger()
        inclure_archivees = self.var_inclure_archivees.get()
        factures = [
            f for f in self._factures_brutes
            if (inclure_archivees or f["id"] not in archivees)
            and (not termes_client or correspond(f["client_nom"], termes_client)
                or correspond(f["destination"], termes_client))
            and (not termes_numero or correspond(str(f["numero"]), termes_numero))
            and (not termes_telephone
                or correspond(f.get("telephone", ""), termes_telephone))
        ]

        self._table.delete(*self._table.get_children())
        self._factures = {f["id"]: f for f in factures}
        for indice, f in enumerate(factures):
            tags = (theme.tag_alternance(indice),)
            if f["id"] in archivees:
                tags += ("archivee",)
            self._table.insert(
                "", "end", iid=str(f["id"]), tags=tags,
                values=(f["numero"], f["date_facture"], f["client_nom"],
                       f["destination"], f.get("telephone", ""), format_fcfa(f["total"])),
            )
        self._statut.configure(
            text=t("client_factures_compte").format(
                n=len(factures), total=len(self._factures_brutes),
                hote=self._app.config.hote, port=self._app.config.port),
            foreground=theme.COULEURS["texte_secondaire"])
        total_filtre = sum(f["total"] for f in factures)
        self._label_total_somme.configure(
            text=f"{t('client_total_periode')} : {format_fcfa(total_filtre)}")

    def _facture_selectionnee(self) -> int | None:
        """Une seule facture sélectionnée — utilisé par l'impression (facture
        et bordereau), qui n'a de sens que pour un document à la fois même
        si le tableau autorise la sélection multiple (pour l'archivage)."""
        selection = self._table.selection()
        if not selection:
            messagebox.showinfo(t("client_selection_titre"), t("client_selection_facture_msg"))
            return None
        if len(selection) > 1:
            messagebox.showinfo(t("client_selection_titre"), t("client_selection_unique_msg"))
            return None
        return int(selection[0])

    def _factures_selectionnees(self) -> list[int]:
        """Une ou plusieurs factures sélectionnées — utilisé par l'archivage
        par lot (sélection multiple ou « Tout sélectionner")."""
        selection = self._table.selection()
        if not selection:
            messagebox.showinfo(t("client_selection_titre"), t("client_selection_facture_msg"))
            return []
        return [int(iid) for iid in selection]

    def _tout_selectionner(self) -> None:
        """Sélectionne toutes les factures actuellement affichées (compte
        tenu des filtres en cours), pour un archivage par lot."""
        self._table.selection_set(self._table.get_children())

    def _archiver_selection(self) -> None:
        """Archive la ou les factures sélectionnées — localement à ce poste
        uniquement, indépendamment de l'archivage de l'application
        principale (voir `app.client.archive_client`)."""
        facture_ids = self._factures_selectionnees()
        if not facture_ids:
            return
        archive_client.archiver_plusieurs(facture_ids)
        self._appliquer_filtres_locaux()

    def _desarchiver_selection(self) -> None:
        facture_ids = self._factures_selectionnees()
        if not facture_ids:
            return
        archive_client.desarchiver_plusieurs(facture_ids)
        self._appliquer_filtres_locaux()

    def _telecharger_et_ouvrir(self, contenu: bytes, prefixe: str, numero: int) -> None:
        dossier = Path(tempfile.gettempdir()) / "promatelas-client"
        dossier.mkdir(parents=True, exist_ok=True)
        chemin = dossier / f"{prefixe}_{numero}_{datetime.now():%H%M%S}.pdf"
        chemin.write_bytes(contenu)
        ouvrir_fichier(chemin)

    def _imprimer_facture(self) -> None:
        facture_id = self._facture_selectionnee()
        if facture_id is None or self._app.api is None:
            return
        try:
            contenu = self._app.api.telecharger_pdf(facture_id)
        except ErreurApiClient as exc:
            messagebox.showerror(t("erreur"), str(exc))
            return
        self._telecharger_et_ouvrir(contenu, "FACTURE", self._factures[facture_id]["numero"])

    def _imprimer_bordereau(self) -> None:
        facture_id = self._facture_selectionnee()
        if facture_id is None or self._app.api is None:
            return
        try:
            contenu = self._app.api.telecharger_bordereau(facture_id)
        except ErreurApiClient as exc:
            messagebox.showerror(t("erreur"), str(exc))
            return
        self._telecharger_et_ouvrir(contenu, "BORDEREAU", self._factures[facture_id]["numero"])


class OngletPaiements(ttk.Frame):
    """Recherche + historique + totaux, écriture (versement) réservée à
    `role_client`. `autoriser_versement=False` (utilisé par
    `ApplicationFacturationEtendue`, rôle `role_facturation` — écriture de
    versement refusée par le serveur, 403) masque la saisie de versement et
    la file hors-ligne associée, cohérent avec `PaiementsView` locale qui est
    déjà lecture seule côté application principale."""

    _DEBOUNCE_MS = 300

    def __init__(self, parent: tk.Widget, app: "ApplicationClient",
                autoriser_versement: bool = True) -> None:
        super().__init__(parent)
        self._app = app
        self._autoriser_versement = autoriser_versement
        self._factures: dict[int, dict] = {}
        self._apres_id: str | None = None
        self._construire()

    def _construire(self) -> None:
        entete = ttk.Frame(self, padding=theme.PAD)
        entete.pack(fill="x")
        ttk.Label(entete, text=t("paie_rechercher")).pack(side="left")
        self.var_recherche = tk.StringVar()
        self.var_recherche.trace_add("write", lambda *_: self._recherche_debouncee())
        ttk.Entry(entete, textvariable=self.var_recherche, width=42).pack(
            side="left", padx=(6, 0))
        ttk.Button(entete, text=t("paie_exporter_rapport"),
                  command=self._exporter_excel).pack(side="right")
        if self._autoriser_versement:
            ttk.Button(entete, text=t("paie_renvoyer_attente"),
                      command=self._renvoyer_hors_ligne).pack(side="right", padx=(0, 8))

        corps = ttk.PanedWindow(self, orient="horizontal")
        corps.pack(fill="both", expand=True, padx=theme.PAD, pady=(0, theme.PAD))

        cadre_liste = ttk.Frame(corps)
        colonnes = ("numero", "date", "client", "destination", "telephone",
                   "total", "verse", "restant")
        self._table = ttk.Treeview(cadre_liste, columns=colonnes, show="headings",
                                   selectmode="browse")
        titres = {"numero": t("fact_no_article"), "date": t("date"), "client": t("fact_client"),
                 "destination": t("paie_destination"), "telephone": t("fact_telephone"),
                 "total": t("paie_total"),
                 "verse": t("paie_verse"), "restant": t("paie_restant")}
        for cle in colonnes:
            self._table.heading(cle, text=titres[cle])
            self._table.column(
                cle, anchor="e" if cle in ("total", "verse", "restant") else "center"
                     if cle in ("numero", "date", "telephone") else "w",
                width=100 if cle in ("total", "verse", "restant") else
                     (90 if cle == "numero" else 90 if cle == "date" else
                     110 if cle == "telephone" else 140 if cle == "client" else 130))
        theme.configurer_lignes_alternees(self._table)
        for tag, couleur in (("restant_du", theme.COULEURS["danger"]),
                            ("restant_solde", theme.COULEURS["succes"]),
                            ("restant_trop_percu", theme.COULEURS["info"])):
            self._table.tag_configure(tag, foreground=couleur)
        self._table.bind("<<TreeviewSelect>>", lambda _e: self._afficher_detail())
        self._table.pack(fill="both", expand=True)
        corps.add(cadre_liste, weight=2)

        cadre_detail = ttk.Frame(corps, padding=(theme.PAD, 0))
        ttk.Label(cadre_detail, text=t("paie_historique_versements"),
                  style="SousTitre.TLabel").pack(anchor="w")
        self._liste_versements = tk.Listbox(
            cadre_detail, height=12, font=("Segoe UI", 11),
            activestyle="none", selectbackground=theme.COULEURS["accent_clair"],
            selectforeground=theme.COULEURS["texte"])
        self._liste_versements.pack(fill="both", expand=True, pady=(4, 8))
        if self._autoriser_versement:
            ttk.Button(cadre_detail, text=t("paie_ajouter_versement"), style="Accent.TButton",
                      command=self._ajouter_versement).pack(fill="x")
        self._statut_detail = ttk.Label(cadre_detail, text="", style="Secondaire.TLabel")
        self._statut_detail.pack(anchor="w", pady=(8, 0))
        corps.add(cadre_detail, weight=1)

        pied = ttk.Frame(self, padding=theme.PAD)
        pied.pack(fill="x")
        self._label_totaux = ttk.Label(
            pied, text="", font=theme.POLICES["total_gros"],
            foreground=theme.COULEURS["accent_sombre"])
        self._label_totaux.pack(side="left")
        self._label_hors_ligne = ttk.Label(pied, text="", foreground=theme.COULEURS["info"])
        if self._autoriser_versement:
            self._label_hors_ligne.pack(side="right")

    def _recherche_debouncee(self) -> None:
        if self._apres_id is not None:
            self.after_cancel(self._apres_id)
        self._apres_id = self.after(self._DEBOUNCE_MS, self.actualiser)

    def actualiser(self) -> None:
        api = self._app.api
        if api is None:
            return
        try:
            factures = api.lister_factures_paiements(recherche=self.var_recherche.get().strip())
            totaux = api.totaux_paiements()
        except ErreurInjoignable:
            self._label_totaux.configure(text=t("client_injoignable"),
                                         foreground=theme.COULEURS["danger"])
            self._maj_indicateur_hors_ligne()
            return
        except ErreurApiClient as exc:
            self._label_totaux.configure(text=str(exc), foreground=theme.COULEURS["danger"])
            return

        # Archivage local (indépendant de celui de l'application principale) :
        # masqué silencieusement ici, sans bascule dédiée — voir l'onglet
        # Historique pour inclure/archiver/désarchiver explicitement.
        archivees = archive_client.charger()
        factures = [f for f in factures if f["id"] not in archivees]

        self._table.delete(*self._table.get_children())
        self._factures = {f["id"]: f for f in factures}
        for indice, f in enumerate(factures):
            tag_couleur = ("restant_du" if f["restant"] > 0
                          else "restant_solde" if f["restant"] == 0 else "restant_trop_percu")
            self._table.insert(
                "", "end", iid=str(f["id"]),
                tags=(theme.tag_alternance(indice), tag_couleur),
                values=(f["numero"], format_date(f["date_facture"]), f["client_nom"],
                       f["destination"], f.get("telephone", ""), format_fcfa(f["total"]),
                       format_fcfa(f["verse"]), format_fcfa(f["restant"])),
            )
        self._label_totaux.configure(
            text=(f"{t('paie_totaux_periode')} — {t('paie_total')} : "
                 f"{format_fcfa(totaux['total_facture'])}  {t('paie_verse')} : "
                 f"{format_fcfa(totaux['total_verse'])}  {t('paie_restant')} : "
                 f"{format_fcfa(totaux['total_restant'])}"),
            foreground=theme.COULEURS["accent_sombre"])
        self._maj_indicateur_hors_ligne()

    def _maj_indicateur_hors_ligne(self) -> None:
        if not self._autoriser_versement:
            return
        n = len(queue_hors_ligne.charger())
        self._label_hors_ligne.configure(
            text=t("paie_hors_ligne").format(n=n) if n else "")

    def _facture_selectionnee_id(self) -> int | None:
        selection = self._table.selection()
        return int(selection[0]) if selection else None

    def _afficher_detail(self) -> None:
        facture_id = self._facture_selectionnee_id()
        self._liste_versements.delete(0, "end")
        if facture_id is None or self._app.api is None:
            return
        try:
            detail = self._app.api.obtenir_facture_paiement(facture_id)
        except ErreurApiClient as exc:
            self._statut_detail.configure(text=str(exc))
            return
        if not detail["versements"]:
            self._liste_versements.insert("end", t("paie_aucun_versement"))
        for indice, v in enumerate(detail["versements"]):
            texte = f"{v['date_versement']}   {format_fcfa(v['montant'])}"
            if v.get("remarque"):
                texte += f"   — {v['remarque']}"
            self._liste_versements.insert("end", texte)
            # Vert pour un versement normal, orange pour une correction
            # (montant négatif) — pas besoin de préciser le poste d'origine.
            couleur = theme.COULEURS["succes"] if v["montant"] >= 0 else theme.COULEURS["info"]
            self._liste_versements.itemconfig(indice, foreground=couleur)
        self._statut_detail.configure(text="")

    def _ajouter_versement(self) -> None:
        facture_id = self._facture_selectionnee_id()
        if facture_id is None:
            messagebox.showinfo(t("client_selection_titre"), t("client_selection_facture_msg"))
            return
        dialogue = DialogueVersement(self)
        if dialogue.resultat is None:
            return
        montant, date_versement, remarque = dialogue.resultat
        try:
            self._app.api.creer_versement(facture_id, montant, date_versement, remarque)
        except ErreurPermission:
            messagebox.showerror(t("client_action_refusee_titre"), t("client_action_refusee_msg"))
            return
        except ErreurInjoignable:
            if messagebox.askyesno(t("client_hors_ligne_titre"), t("paie_enregistrer_hors_ligne")):
                queue_hors_ligne.ajouter(facture_id, montant, date_versement, remarque)
                self._maj_indicateur_hors_ligne()
            return
        except ErreurApiClient as exc:
            messagebox.showerror(t("erreur"), str(exc))
            return
        self.actualiser()
        self._afficher_detail()

    def _renvoyer_hors_ligne(self) -> None:
        if self._app.api is None:
            return
        items = queue_hors_ligne.charger()
        if not items:
            messagebox.showinfo(t("client_versements_attente_titre"),
                               t("client_aucun_versement_attente"))
            return
        reussis = 0
        for index in range(len(items) - 1, -1, -1):
            item = items[index]
            try:
                self._app.api.creer_versement(
                    item.facture_id, item.montant,
                    date.fromisoformat(item.date_versement), item.remarque)
            except ErreurApiClient:
                continue
            queue_hors_ligne.retirer(index)
            reussis += 1
        messagebox.showinfo(
            t("client_versements_attente_titre"),
            t("client_versements_envoyes").format(reussis=reussis, total=len(items)))
        self.actualiser()

    def _exporter_excel(self) -> None:
        """Exporte l'ensemble actuellement filtré en Excel (en-têtes fixes en
        chinois, une ligne par versement avec sa date et son montant — voir
        `app.utils.export_paiements_excel`) et l'ouvre."""
        if not self._factures or self._app.api is None:
            messagebox.showinfo(t("client_selection_titre"), t("aucune_donnee"))
            return
        lignes = []
        for f in self._factures.values():
            try:
                detail = self._app.api.obtenir_facture_paiement(f["id"])
                versements = detail.get("versements", [])
            except ErreurApiClient:
                versements = []
            lignes.append({**f, "versements": versements})
        dossier = Path(tempfile.gettempdir()) / "promatelas-client"
        dossier.mkdir(parents=True, exist_ok=True)
        chemin = dossier / f"RAPPORT_PAIEMENTS_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        generer_rapport_paiements(lignes, chemin)
        ouvrir_fichier(chemin)


class OngletVueClient(ttk.Frame):
    """Synthèse d'un client : total dépensé, versé, restant dû, remises —
    identique à la page « Vue client » de l'application principale, servie
    ici via l'API (`GET /clients/recherche`, `GET /paiements/client/{id}/synthese`,
    `GET /paiements/factures`, `GET /paiements/facture/{id}`, accessibles aux
    deux rôles). Affiche aussi la liste des factures du client sur la période
    filtrée, et le contenu + l'historique des versements de la facture
    sélectionnée. Ce poste gère en plus la remise **annuelle** du client
    sélectionné (la remise par facture reste réservée à l'application
    principale, qui écrit les factures)."""

    def __init__(self, parent: tk.Widget, app: "ApplicationClient") -> None:
        super().__init__(parent)
        self._app = app
        self._resultats: list[dict] = []
        self._client_id: int | None = None
        self._total_facture_annee: int = 0
        self._filtre_actif: dict = {}  # {} = toutes années ; sinon annee=.. ou date_debut/date_fin=..
        self._factures: list[dict] = []
        self._construire()

    def _construire(self) -> None:
        entete = ttk.Frame(self, padding=theme.PAD)
        entete.pack(fill="x")
        ttk.Label(entete, text=t("client_rechercher_client")).pack(side="left")
        self.var_recherche = tk.StringVar()
        self.var_recherche.trace_add("write", lambda *_: self._rechercher())
        ttk.Entry(entete, textvariable=self.var_recherche, width=30).pack(
            side="left", padx=(6, 0))

        filtre = ttk.Frame(self, padding=(theme.PAD, 0))
        filtre.pack(fill="x")
        ttk.Label(filtre, text=t("rem_annee")).pack(side="left")
        annee_courante = date.today().year
        self.var_annee = tk.StringVar(value=t("vue_client_toutes_annees"))
        self._combo_annee = ttk.Combobox(
            filtre, textvariable=self.var_annee, state="readonly", width=18,
            values=[t("vue_client_toutes_annees")]
                  + [str(a) for a in range(annee_courante, annee_courante - 8, -1)],
        )
        self._combo_annee.pack(side="left", padx=(6, 0))
        self._combo_annee.bind("<<ComboboxSelected>>", lambda _e: self._filtrer_par_annee())

        ttk.Label(filtre, text=t("vue_client_periode_precise")).pack(
            side="left", padx=(theme.PAD_L, 4))
        ttk.Label(filtre, text=t("du")).pack(side="left")
        self._champ_du = ChampDate(filtre, largeur=12)
        self._champ_du.pack(side="left", padx=(4, 8))
        self._champ_du.vider()
        ttk.Label(filtre, text=t("au")).pack(side="left")
        self._champ_au = ChampDate(filtre, largeur=12)
        self._champ_au.pack(side="left", padx=(4, 8))
        self._champ_au.vider()
        ttk.Button(filtre, text=t("filtrer"), command=self._filtrer_periode).pack(side="left")
        ttk.Button(filtre, text=t("reinitialiser"),
                  command=self._reinitialiser_periode).pack(side="left", padx=(6, 0))

        self._liste = tk.Listbox(self, height=6)
        self._liste.pack(fill="x", padx=theme.PAD, pady=(0, theme.PAD))
        self._liste.bind("<<ListboxSelect>>", lambda _e: self._selectionner())

        cartes = ttk.Frame(self, padding=theme.PAD)
        cartes.pack(fill="x")
        for i in range(4):
            cartes.columnconfigure(i, weight=1)
        self._carte_depense = KPICard(cartes, t("vue_client_total_depense"))
        self._carte_verse = KPICard(cartes, t("vue_client_total_verse"))
        self._carte_restant = KPICard(cartes, t("vue_client_total_restant"))
        self._carte_remises = KPICard(cartes, t("vue_client_total_remises"))
        for i, carte in enumerate((self._carte_depense, self._carte_verse,
                                   self._carte_restant, self._carte_remises)):
            carte.grid(row=0, column=i, sticky="nsew", padx=(0 if i == 0 else theme.PAD, 0))

        ttk.Button(self, text=t("client_remise_appliquer"), style="Accent.TButton",
                  command=self._appliquer_remise_annuelle).pack(
            anchor="w", padx=theme.PAD, pady=(0, theme.PAD))

        corps = ttk.PanedWindow(self, orient="horizontal")
        corps.pack(fill="both", expand=True, padx=theme.PAD, pady=(0, theme.PAD))

        cadre_factures = ttk.Frame(corps)
        ttk.Label(cadre_factures, text=t("vue_client_factures"),
                  style="SousTitre.TLabel").pack(anchor="w", pady=(0, 4))
        self._table_factures = DataTable(cadre_factures, [
            ("numero", t("ventes_numero"), 80, "center"),
            ("date", t("date"), 100, "center"),
            ("telephone", t("fact_telephone"), 100, "center"),
            ("total", t("total"), 110, "e"),
            ("verse", t("paie_verse"), 110, "e"),
            ("restant", t("paie_restant"), 110, "e"),
        ])
        self._table_factures.pack(fill="both", expand=True)
        for tag, couleur in (("restant_du", theme.COULEURS["danger"]),
                            ("restant_solde", theme.COULEURS["succes"]),
                            ("restant_trop_percu", theme.COULEURS["info"])):
            self._table_factures.tree.tag_configure(tag, foreground=couleur)
        self._table_factures.tree.bind(
            "<<TreeviewSelect>>", lambda _e: self._afficher_detail_facture())
        corps.add(cadre_factures, weight=2)

        cadre_detail = ttk.Frame(corps, padding=(theme.PAD, 0))
        ttk.Label(cadre_detail, text=t("vue_client_contenu_facture"),
                  style="SousTitre.TLabel").pack(anchor="w")
        self._table_contenu = DataTable(cadre_detail, [
            ("quantite", t("fact_quantite"), 70, "center"),
            ("designation", t("fact_designation"), 200, "w"),
            ("pu", t("fact_prix_unitaire"), 90, "e"),
            ("total", t("fact_prix_total"), 100, "e"),
        ], hauteur=6)
        self._table_contenu.pack(fill="both", expand=True, pady=(4, theme.PAD))
        ttk.Label(cadre_detail, text=t("paie_historique_versements"),
                  style="SousTitre.TLabel").pack(anchor="w")
        self._liste_versements = tk.Listbox(cadre_detail, height=8)
        self._liste_versements.pack(fill="both", expand=True, pady=(4, 0))
        corps.add(cadre_detail, weight=1)

        self._statut = ttk.Label(self, text="", style="Secondaire.TLabel",
                                 padding=(theme.PAD, 0))
        self._statut.pack(anchor="w")

    def actualiser(self) -> None:
        self._liste.delete(0, "end")
        self._resultats = []
        self._client_id = None
        self._factures = []
        self._table_factures.vider()
        self._table_contenu.vider()
        self._liste_versements.delete(0, "end")
        for carte in (self._carte_depense, self._carte_verse, self._carte_restant,
                     self._carte_remises):
            carte.definir("—")
        self._statut.configure(text="")

    def _rechercher(self) -> None:
        api = self._app.api
        q = self.var_recherche.get().strip()
        self._liste.delete(0, "end")
        self._resultats = []
        if api is None or not q:
            return
        try:
            self._resultats = api.rechercher_clients(q)
        except ErreurApiClient as exc:
            self._statut.configure(text=str(exc), foreground=theme.COULEURS["danger"])
            return
        for client in self._resultats:
            libelle = client["nom"]
            if client.get("adresse"):
                libelle += f" — {client['adresse']}"
            self._liste.insert("end", libelle)

    def _selectionner(self) -> None:
        selection = self._liste.curselection()
        if not selection or self._app.api is None:
            return
        client = self._resultats[selection[0]]
        self._client_id = client["id"]
        self._rafraichir_synthese()

    def _annee_selectionnee(self) -> int | None:
        """Année choisie dans le filtre période, ou None pour « Toutes »."""
        valeur = self.var_annee.get()
        return int(valeur) if valeur.isdigit() else None

    def _parse_champ_date(self, champ: ttk.Entry) -> date | None:
        texte = champ.get().strip()
        if not texte:
            return None
        try:
            return parse_date_affichage(texte)
        except ValueError:
            messagebox.showerror(t("erreur"), t("date_invalide"))
            return None

    def _filtrer_par_annee(self) -> None:
        """Le sélecteur d'année est prioritaire uniquement si aucune période
        précise (Du/Au) n'est déjà saisie — sinon, réinitialise Du/Au."""
        self._champ_du.delete(0, "end")
        self._champ_au.delete(0, "end")
        self._filtre_actif = {"annee": self._annee_selectionnee()}
        self._rafraichir_synthese()

    def _filtrer_periode(self) -> None:
        """Filtre sur une période précise (Du/Au) — prioritaire sur le
        sélecteur d'année, pour bien définir n'importe quelle période
        (ex. un mois précis) plutôt qu'une année entière."""
        date_debut = self._parse_champ_date(self._champ_du)
        date_fin = self._parse_champ_date(self._champ_au)
        if date_debut is None and date_fin is None:
            return
        self._filtre_actif = {"date_debut": date_debut, "date_fin": date_fin}
        self._rafraichir_synthese()

    def _reinitialiser_periode(self) -> None:
        self._champ_du.delete(0, "end")
        self._champ_au.delete(0, "end")
        self.var_annee.set(t("vue_client_toutes_annees"))
        self._filtre_actif = {}
        self._rafraichir_synthese()

    def _lister_factures_filtrees(self) -> list[dict]:
        """Factures usine du client sélectionné correspondant au filtre de
        période actif (même logique que `synthese_client` : une période
        précise, sinon une année, sinon tout l'historique)."""
        if self._client_id is None or self._app.api is None:
            return []
        if "date_debut" in self._filtre_actif or "date_fin" in self._filtre_actif:
            return self._app.api.lister_factures_paiements(
                client_id=self._client_id,
                date_debut=self._filtre_actif.get("date_debut"),
                date_fin=self._filtre_actif.get("date_fin"),
            )
        resultat = self._app.api.lister_factures_paiements(client_id=self._client_id)
        annee = self._filtre_actif.get("annee")
        if annee is not None:
            resultat = [f for f in resultat if f["date_facture"][:4] == str(annee)]
        return resultat

    def _rafraichir_synthese(self) -> None:
        self._table_contenu.vider()
        self._liste_versements.delete(0, "end")
        if self._client_id is None or self._app.api is None:
            return
        try:
            synthese = self._app.api.synthese_client(self._client_id, **self._filtre_actif)
            self._factures = self._lister_factures_filtrees()
        except ErreurApiClient as exc:
            self._statut.configure(text=str(exc), foreground=theme.COULEURS["danger"])
            return
        self._carte_depense.definir(format_fcfa(synthese["total_facture"]))
        self._carte_verse.definir(format_fcfa(synthese["total_verse"]))
        self._carte_restant.definir(format_fcfa(synthese["total_restant"]))
        self._carte_remises.definir(format_fcfa(synthese["total_remises"]))
        self._total_facture_annee = synthese["total_facture"]
        self._statut.configure(text="")

        self._table_factures.vider()
        for facture in self._factures:
            restant = facture["restant"]
            tag_couleur = ("restant_du" if restant > 0
                          else "restant_solde" if restant == 0
                          else "restant_trop_percu")
            self._table_factures.tree.insert(
                "", "end", iid=str(facture["id"]),
                tags=(theme.tag_alternance(len(self._table_factures.tree.get_children())),
                     tag_couleur),
                values=(facture["numero"], format_date(facture["date_facture"]),
                       facture.get("telephone", ""), format_fcfa(facture["total"]),
                       format_fcfa(facture["verse"]), format_fcfa(restant)),
            )

    def _afficher_detail_facture(self) -> None:
        self._table_contenu.vider()
        self._liste_versements.delete(0, "end")
        iid = self._table_factures.selection_iid()
        if iid is None or self._app.api is None:
            return
        try:
            detail = self._app.api.obtenir_facture_paiement(int(iid))
        except ErreurApiClient as exc:
            self._statut.configure(text=str(exc), foreground=theme.COULEURS["danger"])
            return
        for ligne in detail["lignes"]:
            self._table_contenu.ajouter([
                format_nombre(ligne["quantite"]), ligne["designation"],
                format_fcfa(ligne["prix_unitaire"], False),
                format_fcfa(ligne["total"], False),
            ])
        versements = detail["versements"]
        if not versements:
            self._liste_versements.insert("end", t("paie_aucun_versement"))
        for v in versements:
            texte = (f"{format_date(v['date_versement'])}  "
                    f"{format_fcfa(v['montant'])}  ({v['role_origine'] or 'local'})")
            if v.get("remarque"):
                texte += f"   — {v['remarque']}"
            self._liste_versements.insert("end", texte)

    def _appliquer_remise_annuelle(self) -> None:
        if self._client_id is None:
            messagebox.showinfo(t("client_selection_titre"), t("vue_client_selectionner"))
            return
        annee_defaut = date.today().year
        dialogue = DialogueRemiseAnnuelle(self, annee_defaut, self._total_facture_annee)
        if dialogue.resultat is None:
            return
        annee, ca_annuel, taux, note = dialogue.resultat
        try:
            self._app.api.appliquer_remise_annuelle(self._client_id, annee, ca_annuel, taux, note)
        except ErreurApiClient as exc:
            messagebox.showerror(t("erreur"), str(exc))
            return
        messagebox.showinfo(t("succes"), t("rem_enregistree"))
        self._champ_du.delete(0, "end")
        self._champ_au.delete(0, "end")
        self.var_annee.set(str(annee))
        self._filtre_actif = {"annee": annee}
        self._rafraichir_synthese()


class OngletDashboard(ttk.Frame):
    """Équivalent distant de `app.ui.views.dashboard_view.DashboardView` —
    mêmes KPI/graphiques (réutilise `app.ui.components.graphiques_dashboard`),
    alimentés par `ApiClient.stats_*` plutôt que `StatsService` local."""

    def __init__(self, parent: tk.Widget, app: "ApplicationClient") -> None:
        super().__init__(parent)
        self._app = app
        self._canvases: list = []
        self._construire()

    def _construire(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)
        self.rowconfigure(3, weight=1)

        entete = ttk.Frame(self)
        entete.grid(row=0, column=0, sticky="ew", padx=theme.PAD_L,
                    pady=(theme.PAD_L, theme.PAD))
        ttk.Label(entete, text=t("dash_titre"), style="Titre.TLabel").pack(side="left")

        self.var_granularite = tk.StringVar(value=t("dash_par_mois"))
        self._granularites = {
            t("dash_par_jour"): "jour",
            t("dash_par_semaine"): "semaine",
            t("dash_par_mois"): "mois",
        }
        combo_gran = ttk.Combobox(entete, textvariable=self.var_granularite,
                                  state="readonly", width=10,
                                  values=list(self._granularites))
        combo_gran.pack(side="right")
        combo_gran.bind("<<ComboboxSelected>>", lambda _e: self.actualiser())
        self.var_top = tk.StringVar(value=t("dash_top_quantite"))
        combo_top = ttk.Combobox(entete, textvariable=self.var_top,
                                 state="readonly", width=16,
                                 values=[t("dash_top_quantite"), t("dash_top_ca")])
        combo_top.pack(side="right", padx=theme.PAD_S)
        combo_top.bind("<<ComboboxSelected>>", lambda _e: self.actualiser())

        rangee_kpi = ttk.Frame(self)
        rangee_kpi.grid(row=1, column=0, sticky="ew", padx=theme.PAD_L)
        self._cartes: dict[str, KPICard] = {}
        for indice, (cle, titre) in enumerate([
            ("jour", t("dash_ca_jour")),
            ("mois", t("dash_ca_mois")),
            ("annee", t("dash_ca_annee")),
            ("nb", t("dash_nb_ventes")),
            ("panier", t("dash_panier_moyen")),
        ]):
            rangee_kpi.columnconfigure(indice, weight=1, uniform="kpi")
            carte = KPICard(rangee_kpi, titre)
            carte.grid(row=0, column=indice, sticky="nsew",
                       padx=(0 if indice == 0 else theme.PAD_S, 0))
            self._cartes[cle] = carte

        self._statut = ttk.Label(self, text="", style="Secondaire.TLabel")
        self._statut.grid(row=1, column=0, sticky="e", padx=theme.PAD_L)

        rangee1 = ttk.Frame(self)
        rangee1.grid(row=2, column=0, sticky="nsew", padx=theme.PAD_L, pady=theme.PAD)
        rangee1.columnconfigure(0, weight=3)
        rangee1.columnconfigure(1, weight=2)
        rangee1.rowconfigure(0, weight=1)
        self._cadre_top = graph.cadre_graphique(rangee1, 0, t("dash_top_produits"))
        self._cadre_donut = graph.cadre_graphique(rangee1, 1, t("dash_repartition"))

        rangee2 = ttk.Frame(self)
        rangee2.grid(row=3, column=0, sticky="nsew", padx=theme.PAD_L,
                     pady=(0, theme.PAD_L))
        rangee2.columnconfigure(0, weight=1)
        rangee2.rowconfigure(0, weight=1)
        self._cadre_evolution = graph.cadre_graphique(rangee2, 0, t("dash_evolution"))

    def actualiser(self) -> None:
        api = self._app.api
        if api is None:
            return
        aujourdhui = date.today()
        try:
            kpis = api.stats_kpis(aujourdhui)
            debut_annee = date(aujourdhui.year, 1, 1)
            fin_annee = date(aujourdhui.year, 12, 31)
            par = "quantite" if self.var_top.get() == t("dash_top_quantite") else "ca"
            top = api.stats_top_produits(debut_annee, fin_annee, par=par)
            repartition = api.stats_repartition_gamme(debut_annee, fin_annee)
            granularite = self._granularites.get(self.var_granularite.get(), "mois")
            serie = api.stats_serie_ca(debut_annee, fin_annee, granularite=granularite)
        except ErreurInjoignable:
            self._statut.configure(text=t("client_injoignable"),
                                   foreground=theme.COULEURS["danger"])
            return
        except ErreurAuthentification:
            self._statut.configure(text=t("client_jeton_refuse"),
                                   foreground=theme.COULEURS["danger"])
            return
        except ErreurApiClient as exc:
            self._statut.configure(text=str(exc), foreground=theme.COULEURS["danger"])
            return

        self._statut.configure(text="")
        self._cartes["jour"].definir(format_fcfa(kpis["ca_jour"]["valeur"]),
                                     kpis["ca_jour"]["variation_pct"])
        self._cartes["mois"].definir(format_fcfa(kpis["ca_mois"]["valeur"]),
                                     kpis["ca_mois"]["variation_pct"])
        self._cartes["annee"].definir(format_fcfa(kpis["ca_annee"]["valeur"]),
                                      kpis["ca_annee"]["variation_pct"])
        self._cartes["nb"].definir(str(kpis["nb_ventes"]["valeur"]),
                                   kpis["nb_ventes"]["variation_pct"])
        self._cartes["panier"].definir(format_fcfa(kpis["panier_moyen"]))

        texte_vide = t("dash_aucune_vente")
        graph.dessiner_top_produits(
            self._cadre_top, [(p["libelle"], p["valeur"]) for p in top],
            texte_vide, self._canvases)
        graph.dessiner_repartition_gamme(
            self._cadre_donut, [(p["gamme"], p["ca"]) for p in repartition],
            texte_vide, self._canvases)
        graph.dessiner_evolution_ca(
            self._cadre_evolution, [(p["periode"], p["ca"]) for p in serie],
            texte_vide, self._canvases)


class OngletRemises(ttk.Frame):
    """Équivalent distant de `app.ui.views.remises_view.RemisesView` — tableau
    CA annuel/remise par client, écriture via `ApiClient.appliquer_remise_annuelle`
    (déjà existant, ouvert aux deux rôles)."""

    def __init__(self, parent: tk.Widget, app: "ApplicationClient") -> None:
        super().__init__(parent)
        self._app = app
        self._lignes: list[dict] = []
        self._construire()

    def _construire(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)
        ttk.Label(self, text=t("rem_titre"), style="Titre.TLabel").grid(
            row=0, column=0, sticky="w", padx=theme.PAD_L, pady=(theme.PAD_L, theme.PAD))

        barre = tk.Frame(self, bg=theme.COULEURS["carte"], highlightthickness=1,
                         highlightbackground=theme.COULEURS["bordure"])
        barre.grid(row=1, column=0, sticky="ew", padx=theme.PAD_L)
        tk.Label(barre, text=t("rem_annee"), bg=theme.COULEURS["carte"]).pack(
            side="left", padx=(theme.PAD, 4), pady=theme.PAD)
        annee_courante = date.today().year
        self.var_annee = tk.StringVar(value=str(annee_courante))
        ttk.Combobox(barre, textvariable=self.var_annee, state="readonly", width=8,
                     values=[str(a) for a in range(annee_courante, annee_courante - 8,
                                                   -1)]).pack(side="left")
        ttk.Button(barre, text=t("filtrer"), style="Accent.TButton",
                  command=self.actualiser).pack(side="left", padx=theme.PAD)

        tk.Label(barre, text=t("rem_taux"), bg=theme.COULEURS["carte"]).pack(
            side="left", padx=(theme.PAD_L, 4))
        self.var_taux = tk.StringVar()
        ttk.Entry(barre, textvariable=self.var_taux, width=8).pack(side="left")
        tk.Label(barre, text=t("rem_note"), bg=theme.COULEURS["carte"]).pack(
            side="left", padx=(theme.PAD, 4))
        self.var_note = tk.StringVar()
        ttk.Entry(barre, textvariable=self.var_note, width=22).pack(side="left")
        ttk.Button(barre, text=t("rem_calculer"), style="Accent.TButton",
                  command=self._appliquer_taux).pack(side="left", padx=theme.PAD)

        self._statut = ttk.Label(self, text="", style="Secondaire.TLabel")
        self._statut.grid(row=1, column=0, sticky="e", padx=theme.PAD_L)

        self._table = DataTable(self, [
            ("client", t("fact_client"), 200, "w"),
            ("localite", t("cli_localite"), 130, "w"),
            ("ca", t("rem_ca_annuel"), 150, "e"),
            ("taux", t("rem_taux"), 90, "center"),
            ("montant", t("rem_montant"), 150, "e"),
            ("note", t("rem_note"), 200, "w"),
        ])
        self._table.grid(row=2, column=0, sticky="nsew", padx=theme.PAD_L, pady=theme.PAD)

    def actualiser(self) -> None:
        api = self._app.api
        if api is None:
            return
        try:
            self._lignes = api.remises_tableau(int(self.var_annee.get()))
        except ErreurInjoignable:
            self._statut.configure(text=t("client_injoignable"),
                                   foreground=theme.COULEURS["danger"])
            return
        except ErreurAuthentification:
            self._statut.configure(text=t("client_jeton_refuse"),
                                   foreground=theme.COULEURS["danger"])
            return
        except ErreurApiClient as exc:
            self._statut.configure(text=str(exc), foreground=theme.COULEURS["danger"])
            return

        self._statut.configure(text="")
        self._table.vider()
        for ligne in self._lignes:
            self._table.ajouter([
                ligne["client_nom"], ligne["client_adresse"],
                format_fcfa(ligne["ca_annuel"], False),
                f"{ligne['taux']:g} %" if ligne["taux"] else "",
                format_fcfa(ligne["montant"], False) if ligne["montant"] else "",
                ligne["note"],
            ])

    def _appliquer_taux(self) -> None:
        iid = self._table.selection_iid()
        if iid is None or not self._lignes:
            return
        ligne = self._lignes[self._table.tree.index(iid)]
        try:
            taux = float(self.var_taux.get().replace(",", "."))
            if taux < 0 or taux > 100:
                raise ValueError
        except ValueError:
            messagebox.showerror(t("erreur"), t("nombre_invalide"))
            return
        if ligne["ca_annuel"] <= 0:
            afficher_toast(self, t("rem_aucun_ca"), succes=False)
            return
        try:
            self._app.api.appliquer_remise_annuelle(
                ligne["client_id"], int(self.var_annee.get()), ligne["ca_annuel"],
                taux, self.var_note.get().strip())
        except ErreurApiClient as exc:
            messagebox.showerror(t("erreur"), str(exc))
            return
        afficher_toast(self, t("rem_enregistree"))
        self.actualiser()


def _parser_entier_import(valeur) -> int | None:
    """Convertit une cellule Excel (int/float/texte, éventuellement avec
    espaces ou virgule) en entier — `None` si vide ou illisible. Utilisé
    uniquement par `OngletImportLegacy`, jamais de levée d'exception : une
    cellule imprévisible doit produire une ligne ignorée, pas un plantage."""
    if valeur is None or isinstance(valeur, bool):
        return None
    if isinstance(valeur, (int, float)):
        return int(round(valeur))
    texte = str(valeur).strip().replace(" ", "").replace(",", "")
    if not texte:
        return None
    try:
        return int(float(texte))
    except ValueError:
        return None


def _parser_date_import(valeur) -> date | None:
    """Convertit une cellule Excel (date/datetime déjà typée par openpyxl,
    ou texte JJ/MM/AAAA ou AAAA-MM-JJ) en `date` — `None` si illisible."""
    if valeur is None:
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    texte = str(valeur).strip()
    if not texte:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texte, fmt).date()
        except ValueError:
            continue
    return None


# Colonnes (0-indexées) du résumé de paiements historiques — voir CLAUDE.md,
# section « Import des paiements historiques ». Jamais lues par nom d'en-tête
# (peut être en chinois), toujours par position.
_IMPORT_COL_DATE, _IMPORT_COL_NUMERO, _IMPORT_COL_CLIENT = 0, 1, 2
_IMPORT_COL_MONTANT, _IMPORT_COL_MONTANT_BIS = 6, 7
_IMPORT_COL_VERSEMENT, _IMPORT_COL_SOLDE, _IMPORT_COL_COMMENTAIRE = 8, 9, 10
_IMPORT_LIGNE_ENTETE = 4  # en-tête du tableau ; les données commencent à la ligne suivante


@dataclass
class AnalyseImportLegacy:
    lignes_valides: list[dict] = field(default_factory=list)
    lignes_ignorees: list[str] = field(default_factory=list)
    avertissements: list[str] = field(default_factory=list)


def analyser_feuille_import_legacy(feuille) -> AnalyseImportLegacy:
    """Lit une feuille openpyxl (résumé de paiements historiques) et sépare
    les lignes en valides (prêtes à envoyer au serveur) / ignorées (erreur
    bloquante) / avertissements (envoyées quand même, mais à vérifier
    manuellement). Fonction pure — testable sans Tkinter, voir
    `tests/test_import_legacy_parsing.py`."""
    resultat = AnalyseImportLegacy()
    for indice, ligne in enumerate(
            feuille.iter_rows(min_row=_IMPORT_LIGNE_ENTETE + 1),
            start=_IMPORT_LIGNE_ENTETE + 1):
        valeurs = [cellule.value for cellule in ligne]
        if len(valeurs) <= _IMPORT_COL_COMMENTAIRE:
            valeurs += [None] * (_IMPORT_COL_COMMENTAIRE + 1 - len(valeurs))
        date_brute = valeurs[_IMPORT_COL_DATE]
        numero_brut = valeurs[_IMPORT_COL_NUMERO]
        if date_brute is None and numero_brut is None:
            break  # première ligne totalement vide : fin du tableau

        numero = _parser_entier_import(numero_brut)
        date_facture = _parser_date_import(date_brute)
        client_nom = str(valeurs[_IMPORT_COL_CLIENT] or "").strip()
        montant = _parser_entier_import(valeurs[_IMPORT_COL_MONTANT])
        montant_bis = _parser_entier_import(valeurs[_IMPORT_COL_MONTANT_BIS])
        if not montant:
            montant = montant_bis
        elif montant_bis and montant != montant_bis:
            resultat.avertissements.append(
                t("import_avert_montants_differents").format(indice, montant, montant_bis))
        versement = _parser_entier_import(valeurs[_IMPORT_COL_VERSEMENT]) or 0
        solde_declare = _parser_entier_import(valeurs[_IMPORT_COL_SOLDE])
        if solde_declare is not None and montant is not None \
                and montant - versement != solde_declare:
            resultat.avertissements.append(t("import_avert_solde_incoherent").format(indice))
        commentaire = str(valeurs[_IMPORT_COL_COMMENTAIRE] or "").strip()

        if numero is None or numero <= 0:
            resultat.lignes_ignorees.append(t("import_ignore_numero_invalide").format(indice))
            continue
        if date_facture is None:
            resultat.lignes_ignorees.append(t("import_ignore_date_invalide").format(indice))
            continue
        if not client_nom:
            resultat.lignes_ignorees.append(t("import_ignore_client_vide").format(indice))
            continue
        if not montant or montant <= 0:
            resultat.lignes_ignorees.append(t("import_ignore_montant_invalide").format(indice))
            continue

        resultat.lignes_valides.append({
            "numero": numero, "date_facture": date_facture.isoformat(),
            "client_nom": client_nom, "montant": montant,
            "versement": versement, "commentaire": commentaire,
        })
    return resultat


class OngletImportLegacy(ttk.Frame):
    """Import ponctuel d'un résumé de paiements historiques (fichier Excel de
    l'ancien système) — voir CLAUDE.md, section « Import des paiements
    historiques ». Ne crée que des factures « coquille » (montant total,
    sans détail d'articles) : le fichier est lu et validé entièrement ici
    (côté client — les en-têtes/le texte peuvent être en chinois, donc jamais
    lus par nom, seulement par position), seules les lignes valides sont
    envoyées au serveur en un seul appel."""

    def __init__(self, parent: tk.Widget, app: "ApplicationClient") -> None:
        super().__init__(parent)
        self._app = app
        self._wb = None
        self._lignes_valides: list[dict] = []
        self._rapport_local: list[str] = []
        self._construire()

    def _construire(self) -> None:
        entete = ttk.Frame(self, padding=theme.PAD)
        entete.pack(fill="x")
        ttk.Button(entete, text=t("import_choisir_fichier"),
                  command=self._choisir_fichier).pack(side="left")
        self._label_fichier = ttk.Label(entete, text=t("import_aucun_fichier"),
                                        style="Secondaire.TLabel")
        self._label_fichier.pack(side="left", padx=(8, 0))

        ttk.Label(entete, text=t("import_feuille")).pack(side="left", padx=(24, 4))
        self._combo_feuille = ttk.Combobox(entete, state="readonly", width=20)
        self._combo_feuille.pack(side="left")
        self._combo_feuille.bind("<<ComboboxSelected>>", lambda _e: self._analyser())

        self._bouton_importer = ttk.Button(
            entete, text=t("import_importer"), style="Accent.TButton",
            command=self._importer, state="disabled")
        self._bouton_importer.pack(side="right")

        self._label_resume = ttk.Label(self, text="", padding=(theme.PAD, 4))
        self._label_resume.pack(fill="x")

        cadre_rapport = ttk.Frame(self, padding=(theme.PAD, 0, theme.PAD, theme.PAD))
        cadre_rapport.pack(fill="both", expand=True)
        self._texte_rapport = tk.Text(cadre_rapport, wrap="word", height=20,
                                      font=("Segoe UI", 10))
        self._texte_rapport.pack(fill="both", expand=True, side="left")
        defilement = ttk.Scrollbar(cadre_rapport, orient="vertical",
                                   command=self._texte_rapport.yview)
        self._texte_rapport.configure(yscrollcommand=defilement.set)
        defilement.pack(fill="y", side="right")
        self._texte_rapport.configure(state="disabled")

    def actualiser(self) -> None:
        pass  # import déclenché à la demande, rien à rafraîchir depuis le serveur

    # Lecture du fichier --------------------------------------------------------
    def _choisir_fichier(self) -> None:
        chemin = filedialog.askopenfilename(
            title=t("import_choisir_fichier"),
            filetypes=[("Excel", "*.xlsx *.xlsm"), (t("import_tous_fichiers"), "*.*")])
        if not chemin:
            return
        try:
            self._wb = openpyxl.load_workbook(chemin, data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001 - fichier illisible/corrompu/non Excel
            messagebox.showerror(t("erreur"), str(exc))
            return
        self._label_fichier.configure(text=Path(chemin).name)
        feuilles = self._wb.sheetnames
        self._combo_feuille.configure(values=feuilles)
        indice_defaut = 3 if len(feuilles) > 3 else len(feuilles) - 1
        self._combo_feuille.current(indice_defaut)
        self._analyser()

    def _analyser(self) -> None:
        """Relit entièrement la feuille sélectionnée via `analyser_feuille_import_legacy`
        (fonction pure, testée indépendamment de Tkinter) et met à jour
        l'aperçu/le bouton Importer en conséquence."""
        if self._wb is None or not self._combo_feuille.get():
            return
        feuille = self._wb[self._combo_feuille.get()]
        analyse = analyser_feuille_import_legacy(feuille)
        self._lignes_valides = analyse.lignes_valides
        self._rapport_local = analyse.lignes_ignorees + analyse.avertissements
        self._label_resume.configure(
            text=t("import_resume").format(
                len(self._lignes_valides), len(analyse.lignes_ignorees)))
        self._afficher_rapport(self._rapport_local)
        self._bouton_importer.configure(
            state="normal" if self._lignes_valides else "disabled")

    def _afficher_rapport(self, lignes: list[str]) -> None:
        self._texte_rapport.configure(state="normal")
        self._texte_rapport.delete("1.0", "end")
        self._texte_rapport.insert(
            "end", "\n".join(lignes) if lignes else t("import_rien_a_signaler"))
        self._texte_rapport.configure(state="disabled")

    # Envoi -----------------------------------------------------------------
    def _importer(self) -> None:
        if not self._lignes_valides or self._app.api is None:
            return
        if not messagebox.askyesno(
                t("confirmer"), t("import_confirmation").format(len(self._lignes_valides))):
            return
        try:
            rapport = self._app.api.importer_paiements_legacy(self._lignes_valides)
        except ErreurPermission:
            messagebox.showerror(t("client_action_refusee_titre"), t("client_action_refusee_msg"))
            return
        except ErreurInjoignable:
            messagebox.showerror(t("erreur"), t("client_injoignable"))
            return
        except ErreurApiClient as exc:
            messagebox.showerror(t("erreur"), str(exc))
            return

        lignes_serveur = [
            t("import_serveur_ignoree").format(
                l["numero"], t("import_raison_" + l["raison"]))
            for l in rapport["ignorees"]
        ]
        self._afficher_rapport(self._rapport_local + lignes_serveur)
        self._label_resume.configure(
            text=t("import_resume_final").format(rapport["importees"], len(rapport["ignorees"])))
        afficher_toast(self, t("import_termine").format(rapport["importees"]))
        self._bouton_importer.configure(state="disabled")
        self._app.actualiser()


class OngletServeur(ttk.Frame):
    """Configuration et pilotage de `SMP-Serveur.exe`, directement
    depuis ce poste — pertinent quand ce poste est celui qui héberge
    désormais la base + le serveur (voir CLAUDE.md, section « Machine de
    facturation »). Reprend exactement la logique de `app.ui.views.api_view`
    (mêmes clés i18n `api_*`, déjà traduites fr/en/zh), mais intégrée comme
    onglet plutôt que comme page de l'application principale — ce poste n'a
    ainsi besoin d'installer que `SMP-Client.exe` (+ `SMP-
    Serveur.exe` dans le même dossier, jamais lancé/configuré séparément)."""

    def __init__(self, parent: tk.Widget, app: "ApplicationClient") -> None:
        super().__init__(parent)
        self._app = app
        self._service = app.api_management
        self._pg_conn_sqlite = None  # ouverte à la demande, voir _pg_service
        self._pg_connexion_validee = False  # verrouille la migration tant que False
        self._construire()

    def _carte(self, titre: str) -> tk.Frame:
        cadre = tk.Frame(self._scroll_frame, bg=theme.COULEURS["carte"],
                         highlightthickness=1,
                         highlightbackground=theme.COULEURS["bordure"])
        cadre.pack(fill="x", padx=theme.PAD_L, pady=(theme.PAD_L, theme.PAD))
        ttk.Label(cadre, text=titre, style="SousTitre.TLabel").pack(
            anchor="w", padx=theme.PAD, pady=(theme.PAD, theme.PAD_S))
        return cadre

    def _construire(self) -> None:
        # Défilement : trois cartes (config API, pilotage serveur, migration
        # PostgreSQL) dépassent la hauteur de la fenêtre (1080x680) sur la
        # plupart des écrans — sans ceci, la carte PostgreSQL (ajoutée en
        # dernier) devient invisible, hors de la zone visible, sans aucun
        # moyen d'y accéder.
        canevas = tk.Canvas(self, bg=theme.COULEURS["fond"], highlightthickness=0)
        defilement = ttk.Scrollbar(self, orient="vertical", command=canevas.yview)
        canevas.configure(yscrollcommand=defilement.set)
        canevas.pack(side="left", fill="both", expand=True)
        defilement.pack(side="right", fill="y")

        self._scroll_frame = ttk.Frame(canevas)
        fenetre_id = canevas.create_window((0, 0), window=self._scroll_frame, anchor="nw")
        self._scroll_frame.bind(
            "<Configure>", lambda _e: canevas.configure(scrollregion=canevas.bbox("all")))
        canevas.bind(
            "<Configure>", lambda e: canevas.itemconfig(fenetre_id, width=e.width))

        def _molette(event: tk.Event) -> None:
            canevas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        # Le défilement à la molette n'est actif que survol de CET onglet
        # (bind_all le temps du survol, jamais en permanence) — sinon il
        # interférerait avec le défilement propre des autres onglets
        # (ex. le tableau Treeview de l'onglet Historique).
        canevas.bind("<Enter>", lambda _e: canevas.bind_all("<MouseWheel>", _molette))
        canevas.bind("<Leave>", lambda _e: canevas.unbind_all("<MouseWheel>"))

        ttk.Label(self._scroll_frame, text=t("api_description"), style="Secondaire.TLabel",
                 wraplength=760, justify="left").pack(
            anchor="w", padx=theme.PAD_L, pady=(theme.PAD_L, 0))

        carte_config = self._carte(t("api_config_titre"))
        ligne_hote = ttk.Frame(carte_config, style="Carte.TFrame")
        ligne_hote.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne_hote, text=t("api_hote"), style="Carte.TLabel",
                  width=28).pack(side="left")
        self.var_hote = tk.StringVar()
        ttk.Entry(ligne_hote, textvariable=self.var_hote, width=22).pack(side="left")
        ttk.Label(ligne_hote, text=t("api_port"), style="Carte.TLabel").pack(
            side="left", padx=(theme.PAD, 4))
        self.var_port = tk.StringVar()
        ttk.Entry(ligne_hote, textvariable=self.var_port, width=8).pack(side="left")
        ttk.Label(carte_config, text=t("api_hote_aide"), style="Secondaire.TLabel").pack(
            anchor="w", padx=theme.PAD)

        ligne_jeton_client = ttk.Frame(carte_config, style="Carte.TFrame")
        ligne_jeton_client.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne_jeton_client, text=t("api_jeton_client"), style="Carte.TLabel",
                  width=28).pack(side="left")
        self.var_jeton_client = tk.StringVar()
        ttk.Entry(ligne_jeton_client, textvariable=self.var_jeton_client, width=44).pack(
            side="left", fill="x", expand=True)
        ttk.Button(ligne_jeton_client, text=t("api_generer_jeton_client"),
                  command=self._generer_jeton_client).pack(side="left", padx=(theme.PAD, 0))

        ligne_jeton_principal = ttk.Frame(carte_config, style="Carte.TFrame")
        ligne_jeton_principal.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne_jeton_principal, text=t("api_jeton_principal"), style="Carte.TLabel",
                  width=28).pack(side="left")
        self.var_jeton_principal = tk.StringVar()
        ttk.Entry(ligne_jeton_principal, textvariable=self.var_jeton_principal,
                 width=44).pack(side="left", fill="x", expand=True)
        ttk.Button(ligne_jeton_principal, text=t("api_generer_jeton_principal"),
                  command=self._generer_jeton_principal).pack(side="left", padx=(theme.PAD, 0))

        ligne_jeton_facturation = ttk.Frame(carte_config, style="Carte.TFrame")
        ligne_jeton_facturation.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne_jeton_facturation, text=t("api_jeton_facturation"), style="Carte.TLabel",
                  width=28).pack(side="left")
        self.var_jeton_facturation = tk.StringVar()
        ttk.Entry(ligne_jeton_facturation, textvariable=self.var_jeton_facturation,
                 width=44).pack(side="left", fill="x", expand=True)
        ttk.Button(ligne_jeton_facturation, text=t("api_generer_jeton_facturation"),
                  command=self._generer_jeton_facturation).pack(side="left", padx=(theme.PAD, 0))

        # Jeton « rôle stock » (app Stock, app/stock/) — même schéma que les
        # trois rôles précédents. NB : nécessite que
        # `ApiManagementService.config()`/`enregistrer_config()` (hors
        # périmètre de ce module) exposent aussi `token_stock` ; lecture
        # défensive via `.get(...)` en attendant pour ne pas planter
        # `actualiser()` si ce n'est pas encore le cas.
        ligne_jeton_stock = ttk.Frame(carte_config, style="Carte.TFrame")
        ligne_jeton_stock.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne_jeton_stock, text=t("api_jeton_stock"), style="Carte.TLabel",
                  width=28).pack(side="left")
        self.var_jeton_stock = tk.StringVar()
        ttk.Entry(ligne_jeton_stock, textvariable=self.var_jeton_stock, width=44).pack(
            side="left", fill="x", expand=True)
        ttk.Button(ligne_jeton_stock, text=t("api_generer_jeton_stock"),
                  command=self._generer_jeton_stock).pack(side="left", padx=(theme.PAD, 0))

        ttk.Button(carte_config, text=t("enregistrer"), style="Accent.TButton",
                  command=self._enregistrer_config).pack(
            anchor="e", padx=theme.PAD, pady=(theme.PAD_S, theme.PAD))

        carte_serveur = self._carte(t("api_serveur_titre"))
        ligne_statut = ttk.Frame(carte_serveur, style="Carte.TFrame")
        ligne_statut.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne_statut, text=t("api_statut") + " :", style="Carte.TLabel").pack(
            side="left")
        self._label_statut = ttk.Label(ligne_statut, text="…", style="Carte.TLabel")
        self._label_statut.pack(side="left", padx=(4, theme.PAD_L))
        ttk.Button(ligne_statut, text="↻ " + t("api_actualiser"),
                  command=self._actualiser_statut).pack(side="left")

        ligne_actions = ttk.Frame(carte_serveur, style="Carte.TFrame")
        ligne_actions.pack(fill="x", padx=theme.PAD, pady=(0, theme.PAD_S))
        ttk.Button(ligne_actions, text="▶ " + t("api_demarrer"), style="Accent.TButton",
                  command=self._demarrer).pack(side="left")
        ttk.Button(ligne_actions, text="■ " + t("api_arreter"), style="Danger.TButton",
                  command=self._arreter).pack(side="left", padx=(theme.PAD, 0))

        self.var_demarrage_auto = tk.BooleanVar()
        ttk.Checkbutton(carte_serveur, text=t("api_demarrage_auto"),
                       variable=self.var_demarrage_auto,
                       command=self._basculer_demarrage_auto).pack(
            anchor="w", padx=theme.PAD, pady=(0, theme.PAD))

        self._label_url = ttk.Label(self._scroll_frame, text="", style="Secondaire.TLabel")
        self._label_url.pack(anchor="w", padx=theme.PAD_L)

        self._construire_carte_postgres()

    def _construire_carte_postgres(self) -> None:
        """Interface de connexion PostgreSQL + migration (scaffolding —
        voir app/services/postgres_migration_service.py). La migration reste
        verrouillée tant que « Tester la connexion » n'a pas réussi dans
        cette session (self._pg_connexion_validee) — toute modification des
        champs de connexion invalide ce statut, pour ne jamais migrer vers
        une connexion qui n'a plus été testée."""
        carte = self._carte(t("pg_titre"))
        ttk.Label(carte, text=t("pg_description"), style="Secondaire.TLabel",
                 wraplength=760, justify="left").pack(
            anchor="w", padx=theme.PAD, pady=(0, theme.PAD_S))

        pg = config_postgres.charger()
        self.var_pg_hote = tk.StringVar(value=pg.hote)
        self.var_pg_port = tk.StringVar(value=str(pg.port))
        self.var_pg_base = tk.StringVar(value=pg.base)
        self.var_pg_utilisateur = tk.StringVar(value=pg.utilisateur)
        self.var_pg_mot_de_passe = tk.StringVar(value=pg.mot_de_passe)
        for variable in (self.var_pg_hote, self.var_pg_port, self.var_pg_base,
                        self.var_pg_utilisateur, self.var_pg_mot_de_passe):
            variable.trace_add("write", lambda *_: self._pg_invalider_connexion())

        ligne1 = ttk.Frame(carte, style="Carte.TFrame")
        ligne1.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne1, text=t("pg_hote"), style="Carte.TLabel", width=14).pack(side="left")
        ttk.Entry(ligne1, textvariable=self.var_pg_hote, width=22).pack(side="left")
        ttk.Label(ligne1, text=t("pg_port"), style="Carte.TLabel").pack(
            side="left", padx=(theme.PAD, 4))
        ttk.Entry(ligne1, textvariable=self.var_pg_port, width=8).pack(side="left")

        ligne2 = ttk.Frame(carte, style="Carte.TFrame")
        ligne2.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne2, text=t("pg_base"), style="Carte.TLabel", width=14).pack(side="left")
        ttk.Entry(ligne2, textvariable=self.var_pg_base, width=22).pack(side="left")

        ligne3 = ttk.Frame(carte, style="Carte.TFrame")
        ligne3.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne3, text=t("pg_utilisateur"), style="Carte.TLabel", width=14).pack(
            side="left")
        ttk.Entry(ligne3, textvariable=self.var_pg_utilisateur, width=22).pack(side="left")

        ligne4 = ttk.Frame(carte, style="Carte.TFrame")
        ligne4.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Label(ligne4, text=t("pg_mot_de_passe"), style="Carte.TLabel", width=14).pack(
            side="left")
        ttk.Entry(ligne4, textvariable=self.var_pg_mot_de_passe, width=22, show="•").pack(
            side="left")

        ligne_test = ttk.Frame(carte, style="Carte.TFrame")
        ligne_test.pack(fill="x", padx=theme.PAD, pady=theme.PAD_S)
        ttk.Button(ligne_test, text="🔌 " + t("pg_tester_connexion"),
                  command=self._pg_tester_connexion).pack(side="left")
        self._label_pg_statut = ttk.Label(ligne_test, text="", style="Carte.TLabel")
        self._label_pg_statut.pack(side="left", padx=(theme.PAD, 0))

        ligne_migration = ttk.Frame(carte, style="Carte.TFrame")
        ligne_migration.pack(fill="x", padx=theme.PAD, pady=(0, theme.PAD_S))
        self._bouton_pg_dry_run = ttk.Button(
            ligne_migration, text=t("pg_migration_dry_run"),
            command=lambda: self._pg_migrer(dry_run=True), state="disabled")
        self._bouton_pg_dry_run.pack(side="left")
        self._bouton_pg_reelle = ttk.Button(
            ligne_migration, text=t("pg_migration_reelle"), style="Danger.TButton",
            command=lambda: self._pg_migrer(dry_run=False), state="disabled")
        self._bouton_pg_reelle.pack(side="left", padx=(theme.PAD, 0))

        self._label_pg_resultat = ttk.Label(
            carte, text="", style="Secondaire.TLabel", wraplength=760, justify="left")
        self._label_pg_resultat.pack(anchor="w", padx=theme.PAD, pady=(0, theme.PAD))

        # Bascule effective : rend PostgreSQL réellement actif pour
        # creer_connexion() (config.moteur_bdd) — distincte de la migration
        # ci-dessus, qui ne fait que COPIER les données sans jamais changer
        # ce que l'application lit/écrit au quotidien (voir CLAUDE.md).
        ttk.Separator(carte, orient="horizontal").pack(
            fill="x", padx=theme.PAD, pady=(0, theme.PAD_S))
        ligne_moteur = ttk.Frame(carte, style="Carte.TFrame")
        ligne_moteur.pack(fill="x", padx=theme.PAD, pady=(0, theme.PAD_S))
        ttk.Label(ligne_moteur, text=t("pg_moteur_actuel") + " :",
                 style="Carte.TLabel").pack(side="left")
        self._label_pg_moteur = ttk.Label(ligne_moteur, text="", style="Carte.TLabel")
        self._label_pg_moteur.pack(side="left", padx=(4, 0))

        ligne_bascule = ttk.Frame(carte, style="Carte.TFrame")
        ligne_bascule.pack(fill="x", padx=theme.PAD, pady=(0, theme.PAD))
        self._bouton_pg_activer = ttk.Button(
            ligne_bascule, text=t("pg_activer"), style="Danger.TButton",
            command=self._pg_activer, state="disabled")
        self._bouton_pg_activer.pack(side="left")
        ttk.Button(ligne_bascule, text=t("pg_revenir_sqlite"),
                  command=self._pg_revenir_sqlite).pack(side="left", padx=(theme.PAD, 0))
        self._maj_label_moteur()

    def actualiser(self) -> None:
        config_serveur = self._service.config()
        self.var_hote.set(config_serveur["host"])
        self.var_port.set(str(config_serveur["port"]))
        self.var_jeton_client.set(config_serveur["token_client"])
        self.var_jeton_principal.set(config_serveur["token_principal"])
        self.var_jeton_facturation.set(config_serveur["token_facturation"])
        self.var_jeton_stock.set(config_serveur.get("token_stock", ""))
        try:
            self.var_demarrage_auto.set(self._service.demarrage_auto_actif())
        except Exception:  # noqa: BLE001 — indisponible sur cette machine, pas d'erreur affichée
            self.var_demarrage_auto.set(False)
        self._actualiser_statut()

    def _actualiser_statut(self) -> None:
        en_ligne = self._service.en_ligne()
        self._label_statut.configure(
            text=t("api_statut_actif") if en_ligne else t("api_statut_inactif"),
            foreground=theme.COULEURS["succes"] if en_ligne else theme.COULEURS["danger"],
        )
        config_serveur = self._service.config()
        self._label_url.configure(
            text=f"{t('api_url_test')} http://{config_serveur['host']}:{config_serveur['port']}/health")

    def _generer_jeton_client(self) -> None:
        self.var_jeton_client.set(self._service.generer_jeton())

    def _generer_jeton_principal(self) -> None:
        self.var_jeton_principal.set(self._service.generer_jeton())

    def _generer_jeton_facturation(self) -> None:
        self.var_jeton_facturation.set(self._service.generer_jeton())

    def _generer_jeton_stock(self) -> None:
        self.var_jeton_stock.set(self._service.generer_jeton())

    def _enregistrer_config(self) -> None:
        hote = self.var_hote.get().strip()
        port_texte = self.var_port.get().strip()
        if not hote or not port_texte:
            messagebox.showerror(t("erreur"), t("api_hote_port_requis"))
            return
        try:
            port = int(port_texte)
        except ValueError:
            messagebox.showerror(t("erreur"), t("nombre_invalide"))
            return
        self._service.enregistrer_config(
            hote, port, self.var_jeton_client.get().strip(),
            self.var_jeton_principal.get().strip(),
            self.var_jeton_facturation.get().strip(),
            token_stock=self.var_jeton_stock.get().strip())
        afficher_toast(self, t("api_config_enregistree"))
        self._actualiser_statut()
        self._app.rafraichir_connexion_auto()

    def _demarrer(self) -> None:
        try:
            self._service.demarrer()
        except ErreurGestionApi as exc:
            messagebox.showerror(t("erreur"), str(exc))
            return
        afficher_toast(self, t("api_demarre"))
        self.after(1200, self._actualiser_statut)

    def _arreter(self) -> None:
        try:
            self._service.arreter()
        except ErreurGestionApi as exc:
            messagebox.showerror(t("erreur"), str(exc))
            return
        afficher_toast(self, t("api_arrete"))
        self.after(500, self._actualiser_statut)

    def _basculer_demarrage_auto(self) -> None:
        actif = self.var_demarrage_auto.get()
        try:
            self._service.activer_demarrage_auto(actif)
        except ErreurGestionApi as exc:
            messagebox.showerror(t("erreur"), str(exc))
            self.var_demarrage_auto.set(not actif)
            return
        afficher_toast(self, t("api_demarrage_auto_active") if actif
                       else t("api_demarrage_auto_desactive"))

    # PostgreSQL (scaffolding — voir postgres_migration_service.py) ------------
    def _pg_config_saisie(self) -> ConfigPostgres | None:
        hote = self.var_pg_hote.get().strip()
        base = self.var_pg_base.get().strip()
        utilisateur = self.var_pg_utilisateur.get().strip()
        if not hote or not base or not utilisateur:
            messagebox.showerror(t("erreur"), t("pg_champs_requis"))
            return None
        try:
            port = int(self.var_pg_port.get().strip())
        except ValueError:
            messagebox.showerror(t("erreur"), t("nombre_invalide"))
            return None
        return ConfigPostgres(
            hote=hote, port=port, base=base, utilisateur=utilisateur,
            mot_de_passe=self.var_pg_mot_de_passe.get(),
        )

    def _pg_invalider_connexion(self) -> None:
        """Toute modification d'un champ de connexion invalide le test déjà
        effectué — la migration se reverrouille tant qu'un nouveau test n'a
        pas réussi sur les valeurs actuelles."""
        if not self._pg_connexion_validee:
            return
        self._pg_connexion_validee = False
        self._bouton_pg_dry_run.configure(state="disabled")
        self._bouton_pg_reelle.configure(state="disabled")
        self._bouton_pg_activer.configure(state="disabled")

    def _pg_service(self) -> PostgresMigrationService:
        if self._pg_conn_sqlite is None:
            self._pg_conn_sqlite = creer_connexion()
        return PostgresMigrationService(self._pg_conn_sqlite)

    def _pg_tester_connexion(self) -> None:
        pg_config = self._pg_config_saisie()
        if pg_config is None:
            return
        config_postgres.sauvegarder(config_postgres.ConfigPostgresStockee(
            hote=pg_config.hote, port=pg_config.port, base=pg_config.base,
            utilisateur=pg_config.utilisateur, mot_de_passe=pg_config.mot_de_passe,
        ))
        self._label_pg_statut.configure(text="…", foreground=theme.COULEURS["texte"])
        self.update_idletasks()
        succes, message = self._pg_service().tester_connexion(pg_config)
        self._pg_connexion_validee = succes
        if succes:
            self._label_pg_statut.configure(
                text=t("pg_connexion_ok") + message, foreground=theme.COULEURS["succes"])
            self._bouton_pg_dry_run.configure(state="normal")
            self._bouton_pg_reelle.configure(state="normal")
            self._bouton_pg_activer.configure(state="normal")
        else:
            self._label_pg_statut.configure(
                text=t("pg_connexion_echec") + message, foreground=theme.COULEURS["danger"])
            self._bouton_pg_dry_run.configure(state="disabled")
            self._bouton_pg_reelle.configure(state="disabled")
            self._bouton_pg_activer.configure(state="disabled")

    def _pg_migrer(self, dry_run: bool) -> None:
        if not self._pg_connexion_validee:
            # Filet de sécurité : les boutons sont normalement désactivés
            # tant que ce n'est pas True (voir _construire_carte_postgres) —
            # message explicite si atteint malgré tout.
            messagebox.showerror(t("erreur"), t("pg_connexion_requise"))
            return
        pg_config = self._pg_config_saisie()
        if pg_config is None:
            return
        if not dry_run and not messagebox.askyesno(
            t("confirmer"), t("pg_migration_confirmer")
        ):
            return
        self._label_pg_resultat.configure(text="…", foreground=theme.COULEURS["texte"])
        self.update_idletasks()
        rapport = self._pg_service().migrer(pg_config, dry_run=dry_run)
        if rapport.reussie:
            cle = "pg_migration_resultat_ok" if dry_run else "pg_migration_resultat_ok_reelle"
            self._label_pg_resultat.configure(
                text=t(cle).format(n=rapport.total_lignes, t=len(rapport.tables)),
                foreground=theme.COULEURS["succes"])
        else:
            details = rapport.erreur_globale or "; ".join(
                f"{tab.table}: {tab.erreur}" for tab in rapport.tables if tab.erreur
            ) or "; ".join(rapport.anomalies_integrite)
            self._label_pg_resultat.configure(
                text=f"{t('pg_migration_resultat_echec')} — {details}",
                foreground=theme.COULEURS["danger"])

    def _maj_label_moteur(self) -> None:
        moteur = config.lire_moteur_bdd()
        self._label_pg_moteur.configure(
            text=moteur.upper(),
            foreground=theme.COULEURS["danger"] if moteur == "postgresql"
                      else theme.COULEURS["succes"])

    def _pg_activer(self) -> None:
        """Rend PostgreSQL réellement actif pour `creer_connexion()` — bascule
        distincte de la migration (qui ne fait que copier les données) : voir
        CLAUDE.md, section « Bascule PostgreSQL ». Nécessite un redémarrage
        manuel (creer_connexion() n'est appelée qu'une fois, au démarrage)."""
        if not self._pg_connexion_validee:
            messagebox.showerror(t("erreur"), t("pg_connexion_requise"))
            return
        pg_config = self._pg_config_saisie()
        if pg_config is None:
            return
        if not messagebox.askyesno(t("confirmer"), t("pg_confirmer_activation")):
            return
        config.definir_config_postgres(
            pg_config.hote, pg_config.port, pg_config.base,
            pg_config.utilisateur, pg_config.mot_de_passe)
        config.definir_moteur_bdd("postgresql")
        self._maj_label_moteur()
        messagebox.showinfo(t("succes"), t("pg_active_confirmation"))

    def _pg_revenir_sqlite(self) -> None:
        if not messagebox.askyesno(t("confirmer"), t("pg_confirmer_retour_sqlite")):
            return
        config.definir_moteur_bdd("sqlite")
        self._maj_label_moteur()
        messagebox.showinfo(t("succes"), t("pg_sqlite_confirmation"))


class ApplicationClient(tk.Tk):
    """Fenêtre unique : onglet Historique/Impression (inchangé), onglet
    Paiements (écriture réservée à `role_client`) et onglet Vue client
    (synthèse + remise annuelle, gérée par ce poste)."""

    def __init__(self) -> None:
        super().__init__()
        self.config: ConfigClient = charger()
        translations.definir_langue_memoire(
            self.config.langue if self.config.langue in _LANGUES_CLIENT else "fr")
        self.geometry("1080x680")
        self.minsize(900, 560)
        theme.appliquer_theme(self)
        self.configure(bg=theme.COULEURS["fond"])
        self.report_callback_exception = self._sur_exception

        self.api: ApiClient | AccesDirectDonnees | None = None
        self._conn_directe = None  # sqlite3.Connection, ouverte à la demande — voir _api_pour
        # Onglet Serveur (optionnel) : pertinent uniquement si ce poste
        # héberge lui-même la base + le serveur — voir CLAUDE.md, section
        # « Machine de facturation ». `preparer_dossiers()` crée `data/` (base,
        # settings.json, log du serveur) à côté de cet exécutable au besoin.
        config.preparer_dossiers()
        self.api_management = ApiManagementService()

        self._construire()
        self.after(150, self._demarrage)
        self._demarrer_serveur_si_actif()

    def _demarrer_serveur_si_actif(self) -> None:
        """Démarre `SMP-Serveur.exe` (processus autonome) si ce poste
        l'a déjà configuré comme actif (`api_actif`, voir
        `ApiManagementService.demarrer`/`arreter`) **ou** s'il héberge lui-même
        sa base (jeton `role_client` local — voir `_config_auto_locale`) :
        dans ce second cas, le serveur doit toujours tourner ici par
        définition, sans dépendre de l'historique de `api_actif` (qui ne se
        met à jour qu'au prochain clic explicite sur « Démarrer »/« Arrêter »
        dans l'onglet Serveur — sans ce second déclencheur, un poste déjà
        configuré avant l'ajout de cette persistance ne relancerait jamais
        son serveur tout seul)."""
        if not (config.api_active() or self._config_auto_locale() is not None):
            return
        try:
            self.api_management.demarrer()
        except Exception:  # noqa: BLE001 — ne doit jamais bloquer l'ouverture de la fenêtre
            traceback.print_exc()
            try:
                config.preparer_dossiers()
                with open(config.DOSSIER_DATA / "erreurs.log", "a",
                         encoding="utf-8") as journal:
                    journal.write(traceback.format_exc() + "\n" + "-" * 70 + "\n")
            except OSError:
                pass

    # Erreurs -------------------------------------------------------------
    def _sur_exception(self, exc, val, tb) -> None:
        traceback.print_exception(exc, val, tb)
        messagebox.showerror(t("erreur"), f"{t('erreur_inattendue')}\n\n{val}")

    # Démarrage -----------------------------------------------------------
    def _config_auto_locale(self) -> ConfigClient | None:
        """Si ce poste héberge déjà son propre serveur (onglet Serveur —
        jeton `role_client` déjà généré/enregistré), retourne une
        `ConfigClient` pointant vers lui-même (127.0.0.1) : jamais besoin de
        ressaisir les mêmes hôte/port/jeton une seconde fois via le bouton
        « Connexion ». `None` si aucun serveur local n'est configuré (usage
        « observateur distant » classique — connexion manuelle nécessaire).
        N'est consultée par `_demarrage()` qu'en l'absence de connexion
        manuelle déjà enregistrée — voir son commentaire pour la priorité."""
        jeton_client = config.lire_api_token("role_client")
        if not jeton_client:
            return None
        return ConfigClient(
            hote="127.0.0.1", port=config.lire_api_port(), token=jeton_client,
            machine_id=self.config.machine_id, langue=self.config.langue,
        )

    def _api_direct(self, cfg: ConfigClient) -> AccesDirectDonnees:
        """Accès direct à la base locale (lectures) — voir
        `app.client.acces_direct`. Connexion ouverte une seule fois et
        réutilisée d'une reconnexion à l'autre (même fichier tant que ce poste
        héberge sa propre base)."""
        if self._conn_directe is None:
            self._conn_directe = creer_connexion()
        return AccesDirectDonnees(self._conn_directe, cfg)

    def _demarrage(self) -> None:
        # Une connexion manuelle déjà enregistrée (`client_config.json`, bouton
        # « Connexion ») est un choix explicite et délibéré de l'utilisateur —
        # elle a toujours priorité sur l'auto-connexion locale, même si un
        # jeton `role_client` existe par ailleurs (ex. onglet Serveur testé/
        # configuré une fois sans intention d'en faire un usage réel sur ce
        # poste). Sans cette priorité, un poste observateur distant configuré
        # pour un hôte Tailscale se reconnectait malgré tout à 127.0.0.1 à
        # chaque lancement — la config manuelle semblait « ignorée ».
        if self.config.complete:
            self.api = ApiClient(self.config)
            self.actualiser()
            return
        auto = self._config_auto_locale()
        if auto is not None:
            self.config = auto
            self.api = self._api_direct(auto)
            self.actualiser()
        # Sinon : ni serveur local configuré, ni connexion manuelle
        # enregistrée — aucune fenêtre imposée au démarrage ; chaque onglet
        # affiche son propre état « hors ligne/injoignable » (voir
        # OngletHistorique/OngletPaiements/... .actualiser()) jusqu'à ce que
        # l'utilisateur configure l'un ou l'autre (onglet Serveur, ou bouton
        # « Connexion » pour un serveur distant).

    def rafraichir_connexion_auto(self) -> None:
        """Appelé par `OngletServeur._enregistrer_config` après avoir
        enregistré la config du serveur local : redérive la connexion
        automatiquement plutôt que de laisser l'utilisateur ressaisir les
        mêmes informations via « Connexion »."""
        auto = self._config_auto_locale()
        if auto is not None:
            self.config = auto
            self.api = self._api_direct(auto)
            self.actualiser()

    def _ouvrir_config(self, obligatoire: bool = False) -> None:
        dialogue = FenetreConfig(self, self.config)
        self.wait_window(dialogue)
        if dialogue.resultat is not None:
            self.config = dialogue.resultat
            sauvegarder(self.config)
            self.api = ApiClient(self.config)
            self.actualiser()
        elif obligatoire:
            messagebox.showwarning(
                t("client_connexion_requise"), t("client_connexion_requise_msg"))

    # Langue --------------------------------------------------------------
    def _changer_langue(self, _evenement=None) -> None:
        nom_choisi = self._combo_langue.get()
        for code in _LANGUES_CLIENT:
            if translations.NOMS_LANGUES[code] == nom_choisi:
                translations.definir_langue_memoire(code)
                self.config.langue = code
                sauvegarder(self.config)
                break
        for enfant in self.winfo_children():
            enfant.destroy()
        self._construire()
        self.actualiser()

    # Construction ----------------------------------------------------------
    def _construire(self) -> None:
        self.title(t("client_titre"))
        entete = ttk.Frame(self, padding=theme.PAD)
        entete.pack(fill="x")
        ttk.Label(entete, text=t("client_titre"), style="Titre.TLabel").pack(side="left")
        ttk.Button(entete, text=t("client_connexion"), command=self._ouvrir_config).pack(
            side="right")
        ttk.Button(entete, text=t("client_actualiser"), style="Accent.TButton",
                  command=self.actualiser).pack(side="right", padx=(0, 8))
        self._combo_langue = ttk.Combobox(
            entete, state="readonly", width=10,
            values=[translations.NOMS_LANGUES[code] for code in _LANGUES_CLIENT],
        )
        self._combo_langue.set(translations.NOMS_LANGUES[translations.langue_active()])
        self._combo_langue.pack(side="right", padx=(0, 8))
        self._combo_langue.bind("<<ComboboxSelected>>", self._changer_langue)

        self._onglets = ttk.Notebook(self)
        self._onglets.pack(fill="both", expand=True)
        self.onglet_historique = OngletHistorique(self._onglets, self)
        self.onglet_paiements = OngletPaiements(self._onglets, self)
        self.onglet_vue_client = OngletVueClient(self._onglets, self)
        self.onglet_dashboard = OngletDashboard(self._onglets, self)
        self.onglet_remises = OngletRemises(self._onglets, self)
        self.onglet_import = OngletImportLegacy(self._onglets, self)
        self.onglet_serveur = OngletServeur(self._onglets, self)
        self._onglets.add(self.onglet_historique, text=t("client_onglet_historique"))
        self._onglets.add(self.onglet_paiements, text=t("client_onglet_paiements"))
        self._onglets.add(self.onglet_vue_client, text=t("client_onglet_vue_client"))
        self._onglets.add(self.onglet_dashboard, text=t("dash_titre"))
        self._onglets.add(self.onglet_remises, text=t("rem_titre"))
        self._onglets.add(self.onglet_import, text=t("client_onglet_import"))
        self._onglets.add(self.onglet_serveur, text=t("client_onglet_serveur"))
        self.onglet_serveur.actualiser()

    def actualiser(self) -> None:
        self.onglet_historique.actualiser()
        self.onglet_paiements.actualiser()
        self.onglet_vue_client.actualiser()
        self.onglet_dashboard.actualiser()
        self.onglet_remises.actualiser()
        self.onglet_import.actualiser()
        self.onglet_serveur.actualiser()


def lancer() -> None:
    ApplicationClient().mainloop()
